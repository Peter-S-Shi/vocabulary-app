from __future__ import annotations

import io
import sqlite3
import tempfile
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from sqlite3 import Connection
from typing import Iterator

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from src import db


BACKUP_FORMAT_VERSION = "1"
BACKUP_APP_NAME = "personal_vocabulary_app"
BACKUP_TABLES = (
    "entries",
    "entry_templates",
    "entry_template_fields",
    "entry_field_values",
    "collections",
    "entry_collections",
    "cards",
    "card_revisions",
    "card_revision_entries",
    "entry_change_events",
    "card_review_states",
    "card_review_logs",
    "quiz_sessions",
    "quiz_item_logs",
)


class BackupError(ValueError):
    """A controlled backup generation or preview error."""


@contextmanager
def _connection(connection: Connection | None = None) -> Iterator[Connection]:
    if connection is not None:
        yield connection
        return
    owned = db.get_connection()
    try:
        yield owned
    finally:
        owned.close()


def _table_exists(connection: Connection, table_name: str) -> bool:
    return connection.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table_name,),
    ).fetchone() is not None


def get_database_file_bytes() -> bytes:
    if not db.DB_PATH.exists():
        raise BackupError("Database file not found. No database backup can be created.")
    temporary_path: Path | None = None
    source = destination = None
    try:
        with tempfile.NamedTemporaryFile(prefix="vocab-backup-", suffix=".sqlite3", delete=False) as temporary:
            temporary_path = Path(temporary.name)
        source = db.get_connection()
        destination = sqlite3.connect(temporary_path)
        source.backup(destination)
        destination.close(); destination = None
        source.close(); source = None
        backup_bytes = temporary_path.read_bytes()
        if not backup_bytes:
            raise BackupError("Could not create database backup.")
        return backup_bytes
    except BackupError:
        raise
    except Exception as error:
        raise BackupError("Could not create database backup.") from error
    finally:
        if destination is not None:
            destination.close()
        if source is not None:
            source.close()
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)


def get_backup_table_columns(table_name: str, connection: Connection | None = None) -> list[str]:
    if table_name not in BACKUP_TABLES:
        raise BackupError(f"Table is not allowed for backup: {table_name}")
    with _connection(connection) as conn:
        if not _table_exists(conn, table_name):
            return []
        return [str(row[1]) for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()]


def get_backup_table_rows(table_name: str, connection: Connection | None = None) -> list[dict]:
    if table_name not in BACKUP_TABLES:
        raise BackupError(f"Table is not allowed for backup: {table_name}")
    with _connection(connection) as conn:
        if not _table_exists(conn, table_name):
            return []
        columns = get_backup_table_columns(table_name, conn)
        order_column = "id" if "id" in columns else "rowid"
        return [
            dict(row)
            for row in conn.execute(
                f"SELECT * FROM {table_name} ORDER BY {order_column}"
            ).fetchall()
        ]


def get_backup_summary(connection: Connection | None = None) -> dict:
    with _connection(connection) as conn:
        counts = {}
        for table_name in BACKUP_TABLES:
            if _table_exists(conn, table_name):
                counts[table_name] = int(conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0])
            else:
                counts[table_name] = 0
        return {
            "entries": counts["entries"],
            "collections": counts["collections"],
            "templates": counts["entry_templates"],
            "entry_field_values": counts["entry_field_values"],
            "quiz_sessions": counts["quiz_sessions"],
            "quiz_item_logs": counts["quiz_item_logs"],
            "review_logs": counts["card_review_logs"],
            "tables": counts,
        }


def _metadata_rows(summary: dict, created_at: str) -> list[tuple[str, object]]:
    return [
        ("backup_created_at", created_at),
        ("app_name", BACKUP_APP_NAME),
        ("backup_format_version", BACKUP_FORMAT_VERSION),
        ("database_path_label", "local_vocab_database"),
        ("table_count", len(BACKUP_TABLES)),
        ("entry_count", summary["entries"]),
        ("collection_count", summary["collections"]),
        ("template_count", summary["templates"]),
        ("quiz_session_count", summary["quiz_sessions"]),
        ("quiz_item_log_count", summary["quiz_item_logs"]),
        ("review_log_count", summary["review_logs"]),
    ]


def build_full_backup_workbook_bytes(connection: Connection | None = None) -> bytes:
    try:
        with _connection(connection) as conn:
            summary = get_backup_summary(conn)
            workbook = Workbook()
            workbook.remove(workbook.active)
            for table_name in BACKUP_TABLES:
                if not _table_exists(conn, table_name):
                    continue
                columns = get_backup_table_columns(table_name, conn)
                rows = get_backup_table_rows(table_name, conn)
                worksheet = workbook.create_sheet(table_name)
                worksheet.append(columns)
                for row in rows:
                    worksheet.append(["" if row.get(column) is None else row.get(column) for column in columns])
                worksheet.freeze_panes = "A2"
                for index, column in enumerate(columns, start=1):
                    longest = max([len(column)] + [len(str(row.get(column) or "")) for row in rows[:200]])
                    worksheet.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 10), 50)
            metadata = workbook.create_sheet("backup_metadata")
            metadata.append(["key", "value"])
            for key, value in _metadata_rows(summary, datetime.now(timezone.utc).isoformat(timespec="seconds")):
                metadata.append([key, value])
            metadata.freeze_panes = "A2"
            metadata.column_dimensions["A"].width = 28
            metadata.column_dimensions["B"].width = 40
            output = io.BytesIO()
            workbook.save(output)
            return output.getvalue()
    except BackupError:
        raise
    except Exception as error:
        raise BackupError("Could not create backup workbook.") from error


def validate_backup_metadata(metadata: dict) -> dict:
    warnings = []
    errors = []
    if not metadata:
        warnings.append("This file does not contain backup metadata. It may not be a backup generated by this app.")
    app_name = str(metadata.get("app_name") or "")
    version = str(metadata.get("backup_format_version") or "")
    if app_name and app_name != BACKUP_APP_NAME:
        warnings.append("Backup app name does not match this application.")
    if version and version != BACKUP_FORMAT_VERSION:
        errors.append("Unsupported backup format version.")
    return {"warnings": warnings, "errors": errors}


def preview_backup_workbook(file_bytes: bytes) -> dict:
    result = {"valid_backup": False, "backup_metadata": {}, "sheets": [], "warnings": [], "errors": []}
    if not file_bytes:
        result["errors"].append("Could not read backup workbook.")
        return result
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        result["errors"].append("Could not read backup workbook.")
        return result

    metadata = {}
    if "backup_metadata" in workbook.sheetnames:
        rows = list(workbook["backup_metadata"].iter_rows(values_only=True))
        for row in rows[1:]:
            if row and row[0] not in (None, ""):
                metadata[str(row[0])] = "" if len(row) < 2 or row[1] is None else row[1]
    validation = validate_backup_metadata(metadata)
    result["backup_metadata"] = metadata
    result["warnings"].extend(validation["warnings"])
    result["errors"].extend(validation["errors"])

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        iterator = worksheet.iter_rows(values_only=True)
        headers = next(iterator, ())
        columns = [str(value) for value in headers if value not in (None, "")]
        row_count = sum(1 for row in iterator if any(value not in (None, "") for value in row))
        result["sheets"].append({"sheet_name": sheet_name, "row_count": row_count, "columns": columns})
    for expected in ("entries", "collections", "entry_templates"):
        if expected not in workbook.sheetnames:
            result["warnings"].append(f"Expected backup sheet is missing: {expected}")
    result["valid_backup"] = not result["errors"] and bool(metadata)
    return result


def build_backup_filename(kind: str, extension: str, now: datetime | None = None) -> str:
    timestamp = (now or datetime.now()).strftime("%Y-%m-%d_%H%M%S")
    safe_kind = "database" if kind == "database" else "full"
    suffix = extension.lower().lstrip(".")
    return f"vocabulary_app_{safe_kind}_backup_{timestamp}.{suffix}"

