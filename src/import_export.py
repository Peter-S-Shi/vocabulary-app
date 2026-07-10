from __future__ import annotations

import csv
import io
import re
from contextlib import contextmanager
from datetime import date, datetime, timezone
from sqlite3 import Connection
from typing import Iterator

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from src.db import get_connection


CANONICAL_COLUMNS = [
    "entry_id",
    "language",
    "explanation_language",
    "entry_type",
    "term",
    "meaning",
    "example",
    "notes",
    "tags",
    "source",
    "status",
    "template_id",
    "template_name",
    "template_type",
    "collections",
]
COLLECTION_COLUMNS = [
    "collection_id",
    "collection_name",
    "collection_description",
    "card_size",
    "position",
    "card_number",
]
TIMESTAMP_COLUMNS = ["created_at", "updated_at"]


@contextmanager
def _connection(connection: Connection | None = None) -> Iterator[Connection]:
    if connection is not None:
        yield connection
        return

    owned_connection = get_connection()
    try:
        yield owned_connection
    finally:
        owned_connection.close()


def _dict_rows(connection: Connection, query: str, params: tuple = ()) -> list[dict]:
    return [dict(row) for row in connection.execute(query, params).fetchall()]


def get_exportable_collections(connection: Connection | None = None) -> list[dict]:
    with _connection(connection) as conn:
        return _dict_rows(
            conn,
            """
            SELECT
                collections.id,
                collections.name,
                collections.description,
                collections.card_size,
                collections.is_system,
                collections.system_type,
                COUNT(entry_collections.entry_id) AS entry_count
            FROM collections
            LEFT JOIN entry_collections
              ON entry_collections.collection_id = collections.id
            GROUP BY collections.id
            ORDER BY collections.is_system DESC, collections.name ASC
            """,
        )


def get_exportable_templates(connection: Connection | None = None) -> list[dict]:
    with _connection(connection) as conn:
        return _dict_rows(
            conn,
            """
            SELECT
                templates.id,
                templates.name,
                templates.template_type,
                templates.language,
                templates.is_system,
                COUNT(entries.id) AS entry_count
            FROM entry_templates AS templates
            LEFT JOIN entries ON entries.template_id = templates.id
            GROUP BY templates.id
            ORDER BY templates.is_system DESC, templates.name ASC
            """,
        )


def _collection_names_by_entry(connection: Connection, entry_ids: list[int]) -> dict[int, str]:
    if not entry_ids:
        return {}
    placeholders = ",".join("?" for _ in entry_ids)
    rows = _dict_rows(
        connection,
        f"""
        SELECT links.entry_id, collections.name
        FROM entry_collections AS links
        JOIN collections ON collections.id = links.collection_id
        WHERE links.entry_id IN ({placeholders})
        ORDER BY links.entry_id, collections.name COLLATE NOCASE, collections.id
        """,
        tuple(entry_ids),
    )
    names: dict[int, list[str]] = {}
    for row in rows:
        names.setdefault(int(row["entry_id"]), []).append(str(row["name"] or ""))
    return {entry_id: "; ".join(value for value in values if value) for entry_id, values in names.items()}


def _template_fields_and_values(
    connection: Connection,
    entry_rows: list[dict],
) -> tuple[dict[int, list[str]], dict[int, dict[str, str]], list[str]]:
    template_ids = sorted({int(row["template_id"]) for row in entry_rows if row.get("template_id") is not None})
    entry_ids = [int(row["entry_id"]) for row in entry_rows]
    if not template_ids:
        return {}, {}, []

    template_placeholders = ",".join("?" for _ in template_ids)
    field_rows = _dict_rows(
        connection,
        f"""
        SELECT id, template_id, field_key, display_order
        FROM entry_template_fields
        WHERE template_id IN ({template_placeholders})
        ORDER BY display_order, id, field_key
        """,
        tuple(template_ids),
    )
    fields_by_template: dict[int, list[str]] = {}
    ordered_field_keys: list[str] = []
    seen_field_keys: set[str] = set()
    for field in field_rows:
        template_id = int(field["template_id"])
        field_key = str(field["field_key"])
        fields_by_template.setdefault(template_id, []).append(field_key)
        if field_key not in seen_field_keys:
            ordered_field_keys.append(field_key)
            seen_field_keys.add(field_key)

    values_by_entry: dict[int, dict[str, str]] = {}
    if entry_ids:
        entry_placeholders = ",".join("?" for _ in entry_ids)
        value_rows = _dict_rows(
            connection,
            f"""
            SELECT values_table.entry_id, fields.field_key, values_table.field_value
            FROM entry_field_values AS values_table
            JOIN entry_template_fields AS fields ON fields.id = values_table.field_id
            WHERE values_table.entry_id IN ({entry_placeholders})
            ORDER BY values_table.entry_id, fields.display_order, fields.id
            """,
            tuple(entry_ids),
        )
        for value in value_rows:
            values_by_entry.setdefault(int(value["entry_id"]), {})[str(value["field_key"])] = value.get("field_value") or ""

    return fields_by_template, values_by_entry, ordered_field_keys


def _base_entry_rows(
    connection: Connection,
    where_sql: str = "",
    params: tuple = (),
    order_sql: str = "entries.id ASC",
) -> list[dict]:
    return _dict_rows(
        connection,
        f"""
        SELECT
            entries.id AS entry_id,
            entries.language,
            entries.explanation_language,
            entries.entry_type,
            entries.term,
            entries.meaning,
            entries.example,
            entries.notes,
            entries.tags,
            entries.source,
            entries.status,
            entries.template_id,
            templates.name AS template_name,
            templates.template_type,
            entries.created_at,
            entries.updated_at
        FROM entries
        LEFT JOIN entry_templates AS templates ON templates.id = entries.template_id
        {where_sql}
        ORDER BY {order_sql}
        """,
        params,
    )


def _flatten_entry_rows(
    connection: Connection,
    entry_rows: list[dict],
    include_template_fields: bool,
) -> list[dict]:
    collection_names = _collection_names_by_entry(connection, [int(row["entry_id"]) for row in entry_rows])
    if include_template_fields:
        fields_by_template, values_by_entry, _ = _template_fields_and_values(connection, entry_rows)
    else:
        fields_by_template, values_by_entry = {}, {}

    flattened = []
    for entry in entry_rows:
        entry_id = int(entry["entry_id"])
        row = {column: entry.get(column) for column in CANONICAL_COLUMNS if column != "collections"}
        row["collections"] = collection_names.get(entry_id, "")
        row.update({column: entry.get(column) for column in TIMESTAMP_COLUMNS})
        if include_template_fields and entry.get("template_id") is not None:
            for field_key in fields_by_template.get(int(entry["template_id"]), []):
                row[f"field:{field_key}"] = values_by_entry.get(entry_id, {}).get(field_key, "")
        flattened.append(row)
    return flattened


def export_all_entries_to_rows(
    include_template_fields: bool = True,
    connection: Connection | None = None,
) -> list[dict]:
    with _connection(connection) as conn:
        entries = _base_entry_rows(conn)
        return _flatten_entry_rows(conn, entries, include_template_fields)


def export_collection_entries_to_rows(
    collection_id: int,
    include_template_fields: bool = True,
    connection: Connection | None = None,
) -> list[dict]:
    with _connection(connection) as conn:
        collection_row = conn.execute(
            """
            SELECT id, name, description, card_size
            FROM collections
            WHERE id = ?
            """,
            (int(collection_id),),
        ).fetchone()
        if collection_row is None:
            return []
        collection = dict(collection_row)
        entries = _base_entry_rows(
            conn,
            "JOIN entry_collections AS target_link ON target_link.entry_id = entries.id WHERE target_link.collection_id = ?",
            (int(collection_id),),
            "target_link.position ASC, target_link.id ASC",
        )
        position_rows = _dict_rows(
            conn,
            """
            SELECT entry_id, position
            FROM entry_collections
            WHERE collection_id = ?
            ORDER BY position, id
            """,
            (int(collection_id),),
        )
        positions = {int(row["entry_id"]): int(row["position"]) for row in position_rows}
        card_size = max(int(collection.get("card_size") or 1), 1)
        flattened = _flatten_entry_rows(conn, entries, include_template_fields)
        for row in flattened:
            position = positions[int(row["entry_id"])]
            row.update(
                {
                    "collection_id": int(collection["id"]),
                    "collection_name": collection.get("name") or "",
                    "collection_description": collection.get("description") or "",
                    "card_size": card_size,
                    "position": position,
                    "card_number": ((position - 1) // card_size) + 1,
                }
            )
        return flattened


def export_template_entries_to_rows(
    template_id: int,
    include_template_fields: bool = True,
    connection: Connection | None = None,
) -> list[dict]:
    with _connection(connection) as conn:
        entries = _base_entry_rows(
            conn,
            "WHERE entries.template_id = ?",
            (int(template_id),),
        )
        return _flatten_entry_rows(conn, entries, include_template_fields)


def get_export_columns(rows: list[dict]) -> list[str]:
    if not rows:
        return [*CANONICAL_COLUMNS, *TIMESTAMP_COLUMNS]

    present = {column for row in rows for column in row}
    ordered = [column for column in CANONICAL_COLUMNS if column in present]
    ordered.extend(column for column in COLLECTION_COLUMNS if column in present)
    ordered.extend(column for column in TIMESTAMP_COLUMNS if column in present)

    dynamic_columns = []
    seen = set(ordered)
    for row in rows:
        for column in row:
            if column.startswith("field:") and column not in seen:
                dynamic_columns.append(column)
                seen.add(column)
    ordered.extend(dynamic_columns)
    ordered.extend(sorted(present - set(ordered)))
    return ordered


def _cell_value(value):
    return "" if value is None else value


def rows_to_csv_bytes(rows: list[dict], columns: list[str] | None = None) -> bytes:
    export_columns = list(columns) if columns is not None else get_export_columns(rows)
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=export_columns, extrasaction="ignore", lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow({column: _cell_value(row.get(column)) for column in export_columns})
    return output.getvalue().encode("utf-8-sig")


def rows_to_xlsx_bytes(
    rows: list[dict],
    columns: list[str] | None = None,
    sheet_name: str = "entries",
) -> bytes:
    export_columns = list(columns) if columns is not None else get_export_columns(rows)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sanitize_sheet_name(sheet_name)
    worksheet.append(export_columns)
    for row in rows:
        worksheet.append([_cell_value(row.get(column)) for column in export_columns])
    worksheet.freeze_panes = "A2"
    for index, column in enumerate(export_columns, start=1):
        longest = max([len(str(column))] + [len(str(_cell_value(row.get(column)))) for row in rows[:200]])
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 10), 50)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def sanitize_filename_component(value: str, fallback: str = "export") -> str:
    safe = re.sub(r"[^a-zA-Z0-9_-]+", "_", str(value or "").strip().lower()).strip("_")
    return safe or fallback


def sanitize_sheet_name(value: str) -> str:
    safe = re.sub(r"[\\/*?:\[\]]+", "_", str(value or "entries")).strip()[:31]
    return safe or "entries"


def build_export_filename(scope: str, label: str, extension: str, export_date: date | None = None) -> str:
    safe_scope = sanitize_filename_component(scope)
    safe_label = sanitize_filename_component(label)
    safe_extension = extension.lower().lstrip(".")
    date_text = (export_date or date.today()).isoformat()
    if safe_scope == "all_entries":
        return f"vocabulary_export_all_entries_{date_text}.{safe_extension}"
    if safe_scope == "collections" and safe_label == "summary":
        return f"vocabulary_collection_summary_{date_text}.{safe_extension}"
    return f"vocabulary_export_{safe_scope}_{safe_label}_{date_text}.{safe_extension}"

IMPORT_COLUMNS = {
    "entry_id",
    "language",
    "explanation_language",
    "entry_type",
    "term",
    "meaning",
    "example",
    "notes",
    "tags",
    "source",
    "status",
    "template_id",
    "template_name",
    "template_type",
    "collections",
    "created_at",
    "updated_at",
    "collection_id",
    "collection_name",
    "collection_description",
    "card_size",
    "position",
    "card_number",
}
KNOWN_STATUSES = {"new", "learning", "familiar", "mastered"}
SUPPORTED_IMPORT_MODES = {"auto", "general_entry", "template_aware", "collection"}
INTEGER_IMPORT_COLUMNS = ("position", "card_number", "card_size")


class ImportPreviewError(ValueError):
    """A controlled file parsing or preview error suitable for UI display."""


def detect_file_type(filename: str, mime_type: str | None = None) -> str:
    del mime_type
    suffix = str(filename or "").strip().lower()
    if suffix.endswith(".csv"):
        return "csv"
    if suffix.endswith(".xlsx"):
        return "xlsx"
    if suffix.endswith(".xls"):
        raise ImportPreviewError("Legacy .xls files are not supported. Please save as .xlsx or .csv.")
    raise ImportPreviewError("Unsupported file format. Please upload .csv or .xlsx.")


def parse_csv_bytes(file_bytes: bytes, encoding: str = "utf-8-sig") -> list[dict]:
    if not file_bytes:
        raise ImportPreviewError("File contains no rows.")
    try:
        content = file_bytes.decode(encoding)
    except (UnicodeDecodeError, LookupError) as error:
        raise ImportPreviewError("Could not decode CSV file as UTF-8.") from error
    try:
        reader = csv.DictReader(io.StringIO(content, newline=""), strict=True)
        if reader.fieldnames is None or not any(str(header or "").strip() for header in reader.fieldnames):
            raise ImportPreviewError("File is missing a header row.")
        normalized_headers = [str(header or "").strip().casefold() for header in reader.fieldnames if str(header or "").strip()]
        if len(normalized_headers) != len(set(normalized_headers)):
            raise ImportPreviewError("File contains duplicate column headers.")
        rows = [dict(row) for row in reader]
        if any(None in row for row in rows):
            raise ImportPreviewError("Could not parse CSV file: a row has more values than the header.")
    except (csv.Error, TypeError) as error:
        raise ImportPreviewError("Could not parse CSV file.") from error
    if not rows:
        raise ImportPreviewError("File contains no rows.")
    return rows


def parse_xlsx_bytes(file_bytes: bytes, sheet_name: str | None = None) -> list[dict]:
    if not file_bytes:
        raise ImportPreviewError("File contains no rows.")
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        if sheet_name is not None:
            if sheet_name not in workbook.sheetnames:
                raise ImportPreviewError(f"Worksheet not found: {sheet_name}")
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active
        values = worksheet.iter_rows(values_only=True)
        headers = next(values, None)
        if headers is None or not any(str(header or "").strip() for header in headers):
            raise ImportPreviewError("File is missing a header row.")
        header_values = ["" if header is None else str(header) for header in headers]
        normalized_headers = [header.strip().casefold() for header in header_values if header.strip()]
        if len(normalized_headers) != len(set(normalized_headers)):
            raise ImportPreviewError("File contains duplicate column headers.")
        rows = [
            {header_values[index]: "" if value is None else value for index, value in enumerate(row)}
            for row in values
        ]
    except ImportPreviewError:
        raise
    except Exception as error:
        raise ImportPreviewError("Could not parse XLSX file.") from error
    if not rows:
        raise ImportPreviewError("File contains no rows.")
    return rows


def get_xlsx_sheet_names(file_bytes: bytes) -> list[str]:
    if not file_bytes:
        raise ImportPreviewError("File contains no rows.")
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        return list(workbook.sheetnames)
    except Exception as error:
        raise ImportPreviewError("Could not read XLSX worksheets.") from error


def _normalized_column_name(column_name) -> str:
    name = str(column_name or "").strip()
    lower_name = name.lower()
    if lower_name in IMPORT_COLUMNS:
        return lower_name
    if lower_name.startswith("field:"):
        return f"field:{name.split(':', 1)[1].strip()}"
    return name


def normalize_import_rows(raw_rows: list[dict]) -> list[dict]:
    normalized = []
    for row_number, raw_row in enumerate(raw_rows, start=2):
        data = {}
        for column, value in dict(raw_row or {}).items():
            normalized_column = _normalized_column_name(column)
            if not normalized_column:
                continue
            normalized_value = "" if value is None else value.strip() if isinstance(value, str) else value
            data[normalized_column] = normalized_value
        if not any(str(value or "").strip() for value in data.values()):
            continue
        normalized.append({"row_number": row_number, "data": data, "raw": dict(raw_row or {})})
    return normalized


def get_template_lookup(connection: Connection | None = None) -> dict:
    with _connection(connection) as conn:
        templates = _dict_rows(
            conn,
            """
            SELECT id, name, template_type, language, is_system
            FROM entry_templates
            ORDER BY id
            """,
        )
    return {
        "templates": templates,
        "by_id": {int(template["id"]): template for template in templates},
        "by_name": {str(template["name"] or "").strip().casefold(): template for template in templates},
        "by_type": {str(template["template_type"] or "").strip().casefold(): template for template in templates},
    }


def get_template_field_lookup(template_id: int, connection: Connection | None = None) -> dict[str, dict]:
    with _connection(connection) as conn:
        rows = _dict_rows(
            conn,
            """
            SELECT id, template_id, field_key, field_label, field_type, required, display_order
            FROM entry_template_fields
            WHERE template_id = ?
            ORDER BY display_order, id
            """,
            (int(template_id),),
        )
    return {str(row["field_key"]): row for row in rows}


def get_collection_lookup(connection: Connection | None = None) -> dict:
    with _connection(connection) as conn:
        rows = _dict_rows(
            conn,
            """
            SELECT id, name, is_system, system_type
            FROM collections
            ORDER BY name
            """,
        )
    return {
        "collections": rows,
        "by_id": {int(row["id"]): row for row in rows},
        "by_name": {str(row["name"] or "").strip().casefold(): row for row in rows},
    }


def _comparison_value(value) -> str:
    return " ".join(str(value or "").strip().casefold().split())


def get_existing_entry_keys(connection: Connection | None = None) -> set[tuple[str, str, str]]:
    with _connection(connection) as conn:
        rows = _dict_rows(conn, "SELECT language, term, meaning FROM entries")
    return {
        (_comparison_value(row.get("language")), _comparison_value(row.get("term")), _comparison_value(row.get("meaning")))
        for row in rows
    }


def _resolve_template(data: dict, template_lookup: dict, mode: str) -> tuple[dict | None, list[str]]:
    errors = []
    name = str(data.get("template_name") or "").strip()
    template_type = str(data.get("template_type") or "").strip()
    template_id = str(data.get("template_id") or "").strip()
    by_name = template_lookup["by_name"]
    by_type = template_lookup["by_type"]
    by_id = template_lookup["by_id"]

    candidates = []
    if name:
        match = by_name.get(name.casefold())
        if match is None:
            errors.append(f"Unknown template: {name}")
        else:
            candidates.append(match)
    if template_type:
        match = by_type.get(template_type.casefold())
        if match is None:
            errors.append(f"Unknown template type: {template_type}")
        else:
            candidates.append(match)
    if template_id:
        try:
            match = by_id.get(int(template_id))
        except ValueError:
            match = None
        if match is None:
            errors.append(f"Unknown template ID: {template_id}")
        else:
            candidates.append(match)

    if candidates and any(candidate["id"] != candidates[0]["id"] for candidate in candidates[1:]):
        errors.append("Template name, type, and ID do not refer to the same template.")
    template = candidates[0] if candidates else None
    if mode == "template_aware" and template is None and not errors:
        errors.append("Template-aware validation requires a known template.")
    if mode == "general_entry" and template is not None and template.get("template_type") != "general":
        errors.append("General Entry mode does not accept a non-general template.")
    return template, errors


def _canonical_values(
    data: dict,
    template: dict | None,
    template_fields: dict[str, dict] | None = None,
) -> tuple[str, str]:
    term = str(data.get("term") or data.get("field:term") or "").strip()
    meaning = str(data.get("meaning") or data.get("field:meaning") or data.get("field:definition") or "").strip()
    template_type = str((template or {}).get("template_type") or data.get("template_type") or "").casefold()
    if not term:
        source_key = {
            "french_verb_present": "infinitive",
            "french_adjective_agreement": "masculine_singular",
            "french_noun_gender_plural": "singular",
            "general": "term",
        }.get(template_type)
        if source_key:
            term = str(data.get(f"field:{source_key}") or "").strip()
    if not term and template_fields:
        ordered_fields = sorted(template_fields.values(), key=lambda field: (int(field.get("display_order") or 0), int(field.get("id") or 0)))
        first_required_text = next(
            (
                field for field in ordered_fields
                if field.get("required") and field.get("field_type") == "text"
            ),
            None,
        )
        if first_required_text is not None:
            term = str(data.get(f"field:{first_required_text['field_key']}") or "").strip()
    return term, meaning


def _split_collection_names(value) -> list[str]:
    return [name.strip() for name in str(value or "").split(";") if name.strip()]


def detect_duplicate_candidates(
    rows: list[dict],
    existing_keys: set[tuple[str, str, str]] | None = None,
) -> set[int]:
    known = set(existing_keys or set())
    duplicates = set()
    seen_import_keys = set()
    for row in rows:
        data = row.get("data", row)
        term = data.get("resolved_term", data.get("term"))
        meaning = data.get("resolved_meaning", data.get("meaning"))
        key = (_comparison_value(data.get("language")), _comparison_value(term), _comparison_value(meaning))
        if all(key) and (key in known or key in seen_import_keys):
            duplicates.add(int(row.get("row_number", 0)))
        if all(key):
            seen_import_keys.add(key)
    return duplicates


def validate_import_rows(
    normalized_rows: list[dict],
    mode: str = "auto",
    options: dict | None = None,
    connection: Connection | None = None,
) -> dict:
    if mode not in SUPPORTED_IMPORT_MODES:
        raise ImportPreviewError(f"Unsupported validation mode: {mode}")
    options = options or {}
    strict_template_fields = bool(options.get("strict_template_fields", True))

    with _connection(connection) as conn:
        template_lookup = get_template_lookup(conn)
        collection_lookup = get_collection_lookup(conn)
        existing_keys = get_existing_entry_keys(conn)
        field_lookup_cache: dict[int, dict[str, dict]] = {}
        prepared_rows = []

        for normalized_row in normalized_rows:
            row_number = int(normalized_row.get("row_number") or 0)
            data = dict(normalized_row.get("data") or {})
            errors = []
            warnings = []
            if not str(data.get("language") or "").strip():
                errors.append("Missing required value: language")

            template, template_errors = _resolve_template(data, template_lookup, mode)
            errors.extend(template_errors)
            if template is None and not template_errors:
                template = template_lookup["by_type"].get("general") or template_lookup["by_name"].get("general entry")

            fields = {}
            if template is not None:
                fields = field_lookup_cache.setdefault(int(template["id"]), get_template_field_lookup(int(template["id"]), conn))
            resolved_term, resolved_meaning = _canonical_values(data, template, fields)
            if not resolved_term:
                errors.append("Canonical term cannot be resolved.")
            if not resolved_meaning:
                errors.append("Canonical meaning cannot be resolved.")
            data["resolved_term"] = resolved_term
            data["resolved_meaning"] = resolved_meaning
            if template is not None:
                data["resolved_template_id"] = int(template["id"])
                data["resolved_template_name"] = template["name"]
                data["resolved_template_type"] = template["template_type"]
                unknown_fields = [column.split(":", 1)[1] for column in data if column.startswith("field:") and column.split(":", 1)[1] not in fields]
                for field_key in unknown_fields:
                    message = f"Unknown template field: field:{field_key}"
                    if strict_template_fields:
                        errors.append(message)
                    else:
                        warnings.append(message)
                has_explicit_template = any(str(data.get(key) or "").strip() for key in ("template_id", "template_name", "template_type"))
                if mode == "template_aware" or (mode == "collection" and has_explicit_template):
                    for field_key, field in fields.items():
                        if field.get("required") and not str(data.get(f"field:{field_key}") or "").strip():
                            errors.append(f"Missing required field: {field_key}")
            elif any(column.startswith("field:") for column in data):
                errors.append("Template field columns require a known template.")

            for collection_name in _split_collection_names(data.get("collections")):
                if collection_name.casefold() not in collection_lookup["by_name"]:
                    warnings.append(f"Unknown collection: {collection_name}")
            for integer_column in INTEGER_IMPORT_COLUMNS:
                value = str(data.get(integer_column) or "").strip()
                if value:
                    try:
                        parsed_value = int(value)
                        if parsed_value < 1:
                            raise ValueError
                    except (TypeError, ValueError):
                        message = f"Invalid positive integer: {integer_column}"
                        if mode == "collection":
                            warnings.append(message)
                        else:
                            errors.append(message)
            status = str(data.get("status") or "").strip()
            if status and status.casefold() not in KNOWN_STATUSES:
                warnings.append("Unknown status value; future import may default to new.")
            for column in data:
                if column not in IMPORT_COLUMNS and not column.startswith("field:") and not column.startswith("resolved_"):
                    warnings.append(f"Unknown column ignored: {column}")

            prepared_rows.append(
                {
                    "row_number": row_number,
                    "data": data,
                    "raw": normalized_row.get("raw", {}),
                    "errors": errors,
                    "warnings": warnings,
                    "duplicate_candidate": False,
                }
            )

        if mode == "collection":
            position_rows: dict[int, list[int]] = {}
            collection_names = set()
            valid_positions = []
            for prepared in prepared_rows:
                data = prepared["data"]
                collection_name = str(data.get("collection_name") or "").strip()
                if collection_name:
                    collection_names.add(collection_name)
                position_text = str(data.get("position") or "").strip()
                card_size_text = str(data.get("card_size") or "").strip()
                card_number_text = str(data.get("card_number") or "").strip()
                try:
                    position = int(position_text) if position_text else None
                    if position is not None and position > 0:
                        valid_positions.append(position)
                        position_rows.setdefault(position, []).append(prepared["row_number"])
                    else:
                        position = None
                except ValueError:
                    position = None
                try:
                    card_size = int(card_size_text) if card_size_text else None
                    if card_size is not None and card_size < 1:
                        card_size = None
                except ValueError:
                    card_size = None
                try:
                    card_number = int(card_number_text) if card_number_text else None
                except ValueError:
                    card_number = None
                if position and card_size and card_number and ((position - 1) // card_size) + 1 != card_number:
                    prepared["warnings"].append("Imported card number is inconsistent and will be recalculated.")
            for position, row_numbers in position_rows.items():
                if len(row_numbers) > 1:
                    for prepared in prepared_rows:
                        if prepared["row_number"] in row_numbers:
                            prepared["warnings"].append(f"Duplicate file position: {position}")
            if valid_positions:
                expected = set(range(min(valid_positions), max(valid_positions) + 1))
                if set(valid_positions) != expected:
                    for prepared in prepared_rows:
                        prepared["warnings"].append("File positions contain gaps; final positions will be normalized.")
            if len(collection_names) > 1:
                for prepared in prepared_rows:
                    prepared["warnings"].append("This file contains multiple collection names; one selected destination will be used.")

        duplicate_rows = detect_duplicate_candidates(prepared_rows, existing_keys)
        valid_rows = []
        invalid_rows = []
        warning_items = []
        for row in prepared_rows:
            if row["row_number"] in duplicate_rows:
                row["duplicate_candidate"] = True
                row["warnings"].append("Possible duplicate entry")
            for warning in row["warnings"]:
                warning_items.append({"row_number": row["row_number"], "message": warning})
            if row["errors"]:
                invalid_rows.append(row)
            else:
                valid_rows.append(row)

    return {
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "warnings": warning_items,
        "summary": {
            "total_rows": len(prepared_rows),
            "valid_count": len(valid_rows),
            "invalid_count": len(invalid_rows),
            "warning_count": len(warning_items),
            "duplicate_candidate_count": len(duplicate_rows),
        },
    }


def build_import_preview(
    file_bytes: bytes,
    filename: str,
    mode: str = "auto",
    options: dict | None = None,
    connection: Connection | None = None,
) -> dict:
    file_type = detect_file_type(filename)
    if file_type == "csv":
        raw_rows = parse_csv_bytes(file_bytes)
    else:
        sheet_name = (options or {}).get("sheet_name")
        raw_rows = parse_xlsx_bytes(file_bytes, sheet_name=sheet_name)
    normalized_rows = normalize_import_rows(raw_rows)
    if not normalized_rows:
        raise ImportPreviewError("File contains no rows.")
    result = validate_import_rows(normalized_rows, mode=mode, options=options, connection=connection)
    result["file_type"] = file_type
    result["filename"] = filename
    return result

GENERAL_ENTRY_FIELD_KEYS = ("term", "meaning", "example", "notes", "tags", "source")
DUPLICATE_HANDLING_OPTIONS = {"skip", "import_anyway"}


def _import_text(data: dict, key: str, default: str = "") -> str:
    value = data.get(key)
    if value in (None, ""):
        value = data.get(f"field:{key}", default)
    return str(value or default).strip()


def import_general_entry_rows(
    valid_rows: list[dict],
    duplicate_handling: str = "skip",
    target_collection_id: int | None = None,
    connection: Connection | None = None,
) -> dict:
    if duplicate_handling not in DUPLICATE_HANDLING_OPTIONS:
        raise ValueError("Duplicate handling must be 'skip' or 'import_anyway'.")

    owns_connection = connection is None
    conn = connection or get_connection()
    result = {
        "attempted_count": len(valid_rows),
        "imported_count": 0,
        "skipped_duplicate_count": 0,
        "failed_count": 0,
        "collection_added_count": 0,
        "imported_entry_ids": [],
        "errors": [],
        "warnings": [],
    }

    try:
        template_lookup = get_template_lookup(conn)
        general_template = template_lookup["by_type"].get("general") or template_lookup["by_name"].get("general entry")
        if general_template is None:
            raise ValueError("General Entry template is not available.")
        general_template_id = int(general_template["id"])
        general_fields = get_template_field_lookup(general_template_id, conn)

        target_collection = None
        next_position = None
        if target_collection_id is not None:
            target_collection = conn.execute(
                "SELECT id, name FROM collections WHERE id = ?",
                (int(target_collection_id),),
            ).fetchone()
            if target_collection is None:
                raise ValueError("Target collection does not exist.")
            next_position = int(
                conn.execute(
                    "SELECT COALESCE(MAX(position), 0) + 1 FROM entry_collections WHERE collection_id = ?",
                    (int(target_collection_id),),
                ).fetchone()[0]
            )

        existing_keys = get_existing_entry_keys(conn)
        batch_keys: set[tuple[str, str, str]] = set()
        for index, row in enumerate(valid_rows):
            row_number = int(row.get("row_number") or index + 2)
            savepoint = f"import_row_{index}"
            conn.execute(f"SAVEPOINT {savepoint}")
            try:
                if row.get("errors"):
                    raise ValueError("Row was not valid at preview time.")
                data = dict(row.get("data") or row)
                explicit_template_name = str(data.get("template_name") or "").strip()
                explicit_template_type = str(data.get("template_type") or "").strip()
                resolved_template_id = data.get("resolved_template_id")
                if explicit_template_name and explicit_template_name.casefold() != str(general_template["name"]).casefold():
                    raise ValueError("Only General Entry rows can be imported in Milestone 8.3.")
                if explicit_template_type and explicit_template_type.casefold() != "general":
                    raise ValueError("Only General Entry rows can be imported in Milestone 8.3.")
                if resolved_template_id not in (None, "", general_template_id) and int(resolved_template_id) != general_template_id:
                    raise ValueError("Only General Entry rows can be imported in Milestone 8.3.")

                language = _import_text(data, "language")
                term = str(data.get("resolved_term") or _import_text(data, "term")).strip()
                meaning = str(data.get("resolved_meaning") or _import_text(data, "meaning")).strip()
                if not language:
                    raise ValueError("Missing required value: language")
                if not term:
                    raise ValueError("Canonical term cannot be resolved.")
                if not meaning:
                    raise ValueError("Canonical meaning cannot be resolved.")

                duplicate_key = (_comparison_value(language), _comparison_value(term), _comparison_value(meaning))
                is_duplicate = duplicate_key in existing_keys or duplicate_key in batch_keys
                if is_duplicate and duplicate_handling == "skip":
                    result["skipped_duplicate_count"] += 1
                    result["warnings"].append({"row_number": row_number, "message": "Skipped duplicate entry"})
                    conn.execute(f"RELEASE SAVEPOINT {savepoint}")
                    continue
                if is_duplicate:
                    result["warnings"].append({"row_number": row_number, "message": "Imported duplicate entry by user choice"})

                status = _import_text(data, "status", "new").casefold()
                if status not in KNOWN_STATUSES:
                    status = "new"
                    result["warnings"].append({"row_number": row_number, "message": "Unknown status defaulted to new"})
                template_values = {
                    "term": term,
                    "meaning": meaning,
                    "example": _import_text(data, "example"),
                    "notes": _import_text(data, "notes"),
                    "tags": _import_text(data, "tags"),
                    "source": _import_text(data, "source"),
                }
                now = datetime.now(timezone.utc).isoformat(timespec="seconds")
                cursor = conn.execute(
                    """
                    INSERT INTO entries (
                        template_id, language, explanation_language, entry_type,
                        term, meaning, example, notes, tags, source, status,
                        created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        general_template_id,
                        language,
                        _import_text(data, "explanation_language"),
                        _import_text(data, "entry_type", "word") or "word",
                        term,
                        meaning,
                        template_values["example"],
                        template_values["notes"],
                        template_values["tags"],
                        template_values["source"],
                        status,
                        now,
                        now,
                    ),
                )
                entry_id = int(cursor.lastrowid)
                for field_key in GENERAL_ENTRY_FIELD_KEYS:
                    field = general_fields.get(field_key)
                    if field is None:
                        continue
                    conn.execute(
                        """
                        INSERT INTO entry_field_values (
                            entry_id, field_id, field_value, created_at, updated_at
                        )
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (entry_id, int(field["id"]), template_values[field_key], now, now),
                    )

                if target_collection_id is not None and next_position is not None:
                    conn.execute(
                        """
                        INSERT INTO entry_collections (entry_id, collection_id, position, added_at)
                        VALUES (?, ?, ?, ?)
                        """,
                        (entry_id, int(target_collection_id), next_position, now),
                    )
                    next_position += 1
                    result["collection_added_count"] += 1

                if str(data.get("collections") or "").strip():
                    result["warnings"].append(
                        {"row_number": row_number, "message": "File collections were ignored; only the selected target collection was applied"}
                    )
                result["imported_entry_ids"].append(entry_id)
                result["imported_count"] += 1
                existing_keys.add(duplicate_key)
                batch_keys.add(duplicate_key)
                conn.execute(f"RELEASE SAVEPOINT {savepoint}")
            except Exception as error:
                conn.execute(f"ROLLBACK TO SAVEPOINT {savepoint}")
                conn.execute(f"RELEASE SAVEPOINT {savepoint}")
                result["failed_count"] += 1
                result["errors"].append({"row_number": row_number, "message": str(error)})

        if owns_connection:
            conn.commit()
        return result
    except Exception:
        if owns_connection:
            conn.rollback()
        raise
    finally:
        if owns_connection:
            conn.close()

TEMPLATE_ENTRY_TYPE_DEFAULTS = {
    "french_verb_present": "verb",
    "french_adjective_agreement": "adjective",
    "french_noun_gender_plural": "noun",
}


def get_template_import_template_rows(
    template_id: int,
    blank_row_count: int = 5,
    connection: Connection | None = None,
) -> list[dict]:
    with _connection(connection) as conn:
        template = conn.execute(
            """
            SELECT id, name, template_type, language
            FROM entry_templates
            WHERE id = ?
            """,
            (int(template_id),),
        ).fetchone()
        if template is None:
            raise ValueError("Template not found.")
        fields = _dict_rows(
            conn,
            """
            SELECT field_key
            FROM entry_template_fields
            WHERE template_id = ?
            ORDER BY display_order, id
            """,
            (int(template_id),),
        )

    language = str(template["language"] or "").strip()
    if language.casefold() == "any":
        language = ""
    template_type = str(template["template_type"] or "").strip()
    row_count = max(int(blank_row_count), 1)
    base_row = {
        "language": language,
        "explanation_language": "",
        "entry_type": TEMPLATE_ENTRY_TYPE_DEFAULTS.get(template_type, "word"),
        "term": "",
        "meaning": "",
        "template_name": template["name"],
        "template_type": template_type,
        "status": "new",
        "example": "",
        "notes": "",
        "tags": "",
        "source": "",
    }
    for field in fields:
        base_row[f"field:{field['field_key']}"] = ""

    return [dict(base_row) for _ in range(row_count)]


def build_template_import_template_filename(
    template_name: str,
    extension: str,
) -> str:
    safe_name = sanitize_filename_component(template_name, fallback="template")
    safe_extension = extension.lower().lstrip(".")
    return f"vocabulary_import_template_{safe_name}.{safe_extension}"


def import_template_entry_rows(
    valid_rows: list[dict],
    duplicate_handling: str = "skip",
    target_collection_id: int | None = None,
    connection: Connection | None = None,
) -> dict:
    if duplicate_handling not in DUPLICATE_HANDLING_OPTIONS:
        raise ValueError("Duplicate handling must be 'skip' or 'import_anyway'.")

    owns_connection = connection is None
    conn = connection or get_connection()
    result = {
        "attempted_count": len(valid_rows),
        "imported_count": 0,
        "skipped_duplicate_count": 0,
        "failed_count": 0,
        "field_value_count": 0,
        "added_to_collection_count": 0,
        "imported_entry_ids": [],
        "errors": [],
        "warnings": [],
    }
    try:
        template_lookup = get_template_lookup(conn)
        target_collection = None
        next_position = None
        if target_collection_id is not None:
            target_collection = conn.execute("SELECT id, name FROM collections WHERE id = ?", (int(target_collection_id),)).fetchone()
            if target_collection is None:
                raise ValueError("Target collection does not exist.")
            next_position = int(conn.execute("SELECT COALESCE(MAX(position), 0) + 1 FROM entry_collections WHERE collection_id = ?", (int(target_collection_id),)).fetchone()[0])

        existing_keys = get_existing_entry_keys(conn)
        batch_keys: set[tuple[str, str, str]] = set()
        field_cache: dict[int, dict[str, dict]] = {}
        for index, row in enumerate(valid_rows):
            row_number = int(row.get("row_number") or index + 2)
            savepoint = f"template_import_row_{index}"
            conn.execute(f"SAVEPOINT {savepoint}")
            try:
                if row.get("errors"):
                    raise ValueError("Row was not valid at preview time.")
                data = dict(row.get("data") or row)
                resolved_template_id = data.get("resolved_template_id")
                if resolved_template_id in (None, ""):
                    template, template_errors = _resolve_template(data, template_lookup, "template_aware")
                    if template_errors or template is None:
                        raise ValueError("; ".join(template_errors) or "Template cannot be resolved.")
                else:
                    template = template_lookup["by_id"].get(int(resolved_template_id))
                    if template is None:
                        raise ValueError("Template no longer exists.")
                template_id = int(template["id"])
                fields = field_cache.setdefault(template_id, get_template_field_lookup(template_id, conn))
                unknown_fields = [column for column in data if column.startswith("field:") and column.split(":", 1)[1] not in fields]
                if unknown_fields:
                    raise ValueError(f"Unknown template field: {unknown_fields[0]}")
                for field_key, field in fields.items():
                    if field.get("required") and not str(data.get(f"field:{field_key}") or "").strip():
                        raise ValueError(f"Missing required field: {field_key}")

                language = _import_text(data, "language")
                term, meaning = _canonical_values(data, template, fields)
                if not language:
                    raise ValueError("Missing required value: language")
                if not term:
                    raise ValueError("Canonical term cannot be resolved.")
                if not meaning:
                    raise ValueError("Canonical meaning cannot be resolved.")
                duplicate_key = (_comparison_value(language), _comparison_value(term), _comparison_value(meaning))
                is_duplicate = duplicate_key in existing_keys or duplicate_key in batch_keys
                if is_duplicate and duplicate_handling == "skip":
                    result["skipped_duplicate_count"] += 1
                    result["warnings"].append({"row_number": row_number, "message": "Skipped duplicate entry"})
                    conn.execute(f"RELEASE SAVEPOINT {savepoint}")
                    continue
                if is_duplicate:
                    result["warnings"].append({"row_number": row_number, "message": "Imported duplicate entry by user choice"})

                status = _import_text(data, "status", "new").casefold()
                if status not in KNOWN_STATUSES:
                    status = "new"
                    result["warnings"].append({"row_number": row_number, "message": "Unknown status defaulted to new"})
                template_type = str(template.get("template_type") or "")
                entry_type = _import_text(data, "entry_type", TEMPLATE_ENTRY_TYPE_DEFAULTS.get(template_type, "word")) or TEMPLATE_ENTRY_TYPE_DEFAULTS.get(template_type, "word")
                canonical_optional = {key: _import_text(data, key) for key in ("example", "notes", "tags", "source")}
                now = datetime.now(timezone.utc).isoformat(timespec="seconds")
                cursor = conn.execute(
                    """
                    INSERT INTO entries (
                        template_id, language, explanation_language, entry_type,
                        term, meaning, example, notes, tags, source, status,
                        created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        template_id, language, _import_text(data, "explanation_language"), entry_type,
                        term, meaning, canonical_optional["example"], canonical_optional["notes"],
                        canonical_optional["tags"], canonical_optional["source"], status, now, now,
                    ),
                )
                entry_id = int(cursor.lastrowid)
                row_field_count = 0
                for field_key, field in fields.items():
                    field_value = str(data.get(f"field:{field_key}") or "").strip()
                    if not field_value:
                        continue
                    conn.execute(
                        """
                        INSERT INTO entry_field_values (entry_id, field_id, field_value, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (entry_id, int(field["id"]), field_value, now, now),
                    )
                    row_field_count += 1

                if target_collection_id is not None and next_position is not None:
                    conn.execute(
                        "INSERT INTO entry_collections (entry_id, collection_id, position, added_at) VALUES (?, ?, ?, ?)",
                        (entry_id, int(target_collection_id), next_position, now),
                    )
                    next_position += 1
                    result["added_to_collection_count"] += 1
                if str(data.get("collections") or "").strip():
                    result["warnings"].append({"row_number": row_number, "message": "File collections were ignored; only the selected target collection was applied"})

                result["imported_entry_ids"].append(entry_id)
                result["imported_count"] += 1
                result["field_value_count"] += row_field_count
                existing_keys.add(duplicate_key)
                batch_keys.add(duplicate_key)
                conn.execute(f"RELEASE SAVEPOINT {savepoint}")
            except Exception as error:
                conn.execute(f"ROLLBACK TO SAVEPOINT {savepoint}")
                conn.execute(f"RELEASE SAVEPOINT {savepoint}")
                result["failed_count"] += 1
                result["errors"].append({"row_number": row_number, "message": str(error)})

        if owns_connection:
            conn.commit()
        return result
    except Exception:
        if owns_connection:
            conn.rollback()
        raise
    finally:
        if owns_connection:
            conn.close()


def export_collection_card_entries_to_rows(
    collection_id: int,
    card_number: int,
    include_template_fields: bool = True,
    connection: Connection | None = None,
) -> list[dict]:
    selected_card = int(card_number)
    if selected_card < 1:
        return []
    rows = export_collection_entries_to_rows(collection_id, include_template_fields, connection)
    return [row for row in rows if int(row.get("card_number") or 0) == selected_card]


def export_collections_summary_to_rows(connection: Connection | None = None) -> list[dict]:
    rows = get_exportable_collections(connection)
    return [
        {
            "collection_id": int(row["id"]),
            "collection_name": row.get("name") or "",
            "collection_description": row.get("description") or "",
            "card_size": int(row.get("card_size") or 8),
            "entry_count": int(row.get("entry_count") or 0),
            "card_count": (int(row.get("entry_count") or 0) + max(int(row.get("card_size") or 8), 1) - 1) // max(int(row.get("card_size") or 8), 1),
            "is_system": int(row.get("is_system") or 0),
            "system_type": row.get("system_type") or "",
        }
        for row in rows
    ]


def _collection_import_order(rows: list[dict], preserve_file_order: bool) -> list[dict]:
    if not preserve_file_order:
        return sorted(rows, key=lambda row: int(row.get("row_number") or 0))
    has_valid_position = any(str(row.get("data", {}).get("position") or "").strip().isdigit() and int(row["data"]["position"]) > 0 for row in rows)
    if not has_valid_position:
        return sorted(rows, key=lambda row: int(row.get("row_number") or 0))
    def order_key(row: dict):
        value = str(row.get("data", {}).get("position") or "").strip()
        if value.isdigit() and int(value) > 0:
            return (0, int(value), int(row.get("row_number") or 0))
        return (1, int(row.get("row_number") or 0), 0)
    return sorted(rows, key=order_key)


def import_collection_rows(
    valid_rows: list[dict],
    import_mode: str,
    duplicate_handling: str = "skip",
    target_collection_id: int | None = None,
    create_collection_options: dict | None = None,
    preserve_file_order: bool = True,
    connection: Connection | None = None,
) -> dict:
    if import_mode not in {"append_to_existing", "create_new_collection"}:
        raise ValueError("Unsupported collection import mode.")
    if duplicate_handling not in DUPLICATE_HANDLING_OPTIONS:
        raise ValueError("Duplicate handling must be 'skip' or 'import_anyway'.")
    owns_connection = connection is None
    conn = connection or get_connection()
    result = {
        "attempted_count": len(valid_rows), "imported_entry_count": 0,
        "skipped_duplicate_count": 0, "failed_count": 0,
        "collection_id": None, "collection_name": None,
        "added_to_collection_count": 0, "start_position": None, "end_position": None,
        "created_collection": False, "field_value_count": 0,
        "errors": [], "warnings": [],
    }
    conn.execute("SAVEPOINT collection_import_batch")
    try:
        if import_mode == "append_to_existing":
            if target_collection_id is None:
                raise ValueError("Please select a target collection.")
            collection = conn.execute("SELECT id, name FROM collections WHERE id = ?", (int(target_collection_id),)).fetchone()
            if collection is None:
                raise ValueError("Target collection does not exist.")
            collection_id = int(collection["id"])
            collection_name = collection["name"]
        else:
            options = create_collection_options or {}
            collection_name = str(options.get("name") or "").strip()
            if not collection_name:
                raise ValueError("Please enter a collection name.")
            if conn.execute("SELECT 1 FROM collections WHERE LOWER(name) = LOWER(?)", (collection_name,)).fetchone():
                raise ValueError("A collection with this name already exists.")
            try:
                card_size = int(options.get("card_size") or 8)
            except (TypeError, ValueError) as error:
                raise ValueError("Card size must be a positive integer.") from error
            if card_size < 1:
                raise ValueError("Card size must be a positive integer.")
            now = datetime.now(timezone.utc).isoformat(timespec="seconds")
            cursor = conn.execute(
                "INSERT INTO collections (name, description, card_size, is_system, system_type, created_at, updated_at) VALUES (?, ?, ?, 0, NULL, ?, ?)",
                (collection_name, str(options.get("description") or "").strip(), card_size, now, now),
            )
            collection_id = int(cursor.lastrowid)
            result["created_collection"] = True

        start_max = int(conn.execute("SELECT COALESCE(MAX(position), 0) FROM entry_collections WHERE collection_id = ?", (collection_id,)).fetchone()[0])
        ordered_rows = _collection_import_order(valid_rows, preserve_file_order)
        for row in ordered_rows:
            data = row.get("data", {})
            explicit_template = any(str(data.get(key) or "").strip() for key in ("template_id", "template_name", "template_type"))
            resolved_type = str(data.get("resolved_template_type") or "")
            writer = import_template_entry_rows if explicit_template and resolved_type != "general" else import_general_entry_rows
            row_result = writer([row], duplicate_handling=duplicate_handling, target_collection_id=collection_id, connection=conn)
            result["imported_entry_count"] += row_result["imported_count"]
            result["skipped_duplicate_count"] += row_result["skipped_duplicate_count"]
            result["failed_count"] += row_result["failed_count"]
            result["added_to_collection_count"] += row_result.get("added_to_collection_count", row_result.get("collection_added_count", 0))
            result["field_value_count"] += row_result.get("field_value_count", 0)
            result["errors"].extend(row_result["errors"])
            result["warnings"].extend(row_result["warnings"])

        if result["imported_entry_count"] == 0 and result["created_collection"]:
            conn.execute("ROLLBACK TO SAVEPOINT collection_import_batch")
            result["created_collection"] = False
            result["collection_id"] = None
            result["collection_name"] = None
            result["warnings"].append({"row_number": None, "message": "New collection was not created because no rows were imported"})
        else:
            result["collection_id"] = collection_id
            result["collection_name"] = collection_name
            if result["added_to_collection_count"]:
                result["start_position"] = start_max + 1
                result["end_position"] = start_max + result["added_to_collection_count"]
        conn.execute("RELEASE SAVEPOINT collection_import_batch")
        if owns_connection:
            conn.commit()
        return result
    except Exception:
        conn.execute("ROLLBACK TO SAVEPOINT collection_import_batch")
        conn.execute("RELEASE SAVEPOINT collection_import_batch")
        if owns_connection:
            conn.rollback()
        raise
    finally:
        if owns_connection:
            conn.close()

IMPORT_SAMPLE_LABELS = {
    "general_entry": "General Entry",
    "french_verb_present": "French Verb Present",
    "french_adjective_agreement": "French Adjective Agreement",
    "french_noun_gender_plural": "French Noun Gender Plural",
    "collection_import": "Collection Import",
}


def get_import_sample_rows(sample_key: str) -> list[dict]:
    samples = {
        "general_entry": {
            "language": "English", "explanation_language": "Chinese", "entry_type": "word",
            "term": "subtle", "meaning": "???", "example": "This is a subtle difference.",
            "notes": "", "tags": "academic", "source": "manual", "status": "new",
        },
        "french_verb_present": {
            "language": "French", "explanation_language": "English", "entry_type": "verb",
            "term": "prendre", "meaning": "to take", "template_name": "French Verb Present",
            "template_type": "french_verb_present", "field:infinitive": "prendre", "field:meaning": "to take",
            "field:je": "je prends", "field:tu": "tu prends", "field:il_elle_on": "il/elle/on prend",
            "field:nous": "nous prenons", "field:vous": "vous prenez", "field:ils_elles": "ils/elles prennent",
            "field:example": "Je prends le bus.", "field:notes": "irregular verb",
            "field:tags": "common; irregular", "field:source": "textbook",
        },
        "french_adjective_agreement": {
            "language": "French", "explanation_language": "English", "entry_type": "adjective",
            "term": "bon", "meaning": "good", "template_name": "French Adjective Agreement",
            "template_type": "french_adjective_agreement", "field:masculine_singular": "bon",
            "field:meaning": "good", "field:feminine_singular": "bonne", "field:masculine_plural": "bons",
            "field:feminine_plural": "bonnes", "field:example": "C'est une bonne id?e.",
            "field:notes": "", "field:tags": "agreement", "field:source": "manual",
        },
        "french_noun_gender_plural": {
            "language": "French", "explanation_language": "English", "entry_type": "noun",
            "term": "livre", "meaning": "book", "template_name": "French Noun Gender Plural",
            "template_type": "french_noun_gender_plural", "field:singular": "livre", "field:meaning": "book",
            "field:gender": "masculine", "field:plural": "livres", "field:article": "le",
            "field:example": "Je lis un livre.", "field:notes": "", "field:tags": "noun", "field:source": "manual",
        },
        "collection_import": {
            "collection_name": "My Imported Collection", "collection_description": "Imported vocabulary",
            "card_size": 8, "position": 1, "language": "English", "explanation_language": "Chinese",
            "entry_type": "word", "term": "portable", "meaning": "???", "example": "A portable database.",
            "notes": "", "tags": "sample", "source": "manual", "status": "new",
            "template_name": "General Entry", "template_type": "general",
            "field:term": "portable", "field:meaning": "???", "field:example": "A portable database.",
            "field:notes": "", "field:tags": "sample", "field:source": "manual",
        },
    }
    if sample_key not in samples:
        raise ValueError(f"Unknown import sample: {sample_key}")
    return [samples[sample_key]]


def get_template_field_map_rows(connection: Connection | None = None) -> list[dict]:
    with _connection(connection) as conn:
        return _dict_rows(
            conn,
            """
            SELECT templates.id AS template_id, templates.name AS template_name,
                   templates.template_type, templates.language,
                   fields.field_key, fields.field_label, fields.field_type,
                   fields.required, fields.display_order
            FROM entry_templates AS templates
            LEFT JOIN entry_template_fields AS fields ON fields.template_id = templates.id
            ORDER BY templates.name, fields.display_order, fields.id
            """,
        )

