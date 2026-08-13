from __future__ import annotations

import csv
import io
from pathlib import Path
import sqlite3
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from src import db
from src.backup import BACKUP_TABLES, build_full_backup_workbook_bytes, get_database_file_bytes
from src.collections import create_collection, delete_collection, get_entries_in_collection
from src.entries import delete_entry, get_entry_by_id, update_entry
from src.entry_templates import create_entry_template, create_template_field
from src.import_export import import_general_entry_rows
from src.linked_sources import (
    confirm_collection_source_link,
    confirm_linked_source_refresh,
    get_collection_source_link,
    preview_collection_source_link,
    preview_linked_source_refresh,
    unlink_collection_source,
)
from src.migrations import (
    APP_DATA_VERSION,
    CURRENT_SCHEMA_VERSION,
    LINKED_APPEND_SOURCE_SCHEMA_VERSION,
    MIGRATIONS,
    QUIZ_LOG_HISTORY_SCHEMA_VERSION,
    get_metadata,
    get_schema_version,
    run_migrations,
    set_metadata,
    set_schema_version,
)


CSV_COLUMNS = [
    "language",
    "explanation_language",
    "entry_type",
    "term",
    "meaning",
    "status",
]


class M13BatchBLinkedSourceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory(ignore_cleanup_errors=True)
        self.root = Path(self.temp_dir.name)
        self.original_db_path = db.DB_PATH
        db.DB_PATH = self.root / "m13_batch_b.sqlite3"
        db.init_db()
        self.collection_id = create_collection("Synthetic Linked Collection", card_size=2)

    def tearDown(self) -> None:
        db.DB_PATH = self.original_db_path
        self.temp_dir.cleanup()

    def _row(self, term: str, meaning: str, **overrides) -> dict:
        row = {
            "language": "English",
            "explanation_language": "English",
            "entry_type": "word",
            "term": term,
            "meaning": meaning,
            "status": "new",
        }
        row.update(overrides)
        return row

    def _write_csv(self, rows: list[dict], name: str = "linked-source.csv") -> Path:
        path = self.root / name
        columns = list(dict.fromkeys(CSV_COLUMNS + [key for row in rows for key in row]))
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns)
            writer.writeheader()
            writer.writerows(rows)
        return path

    def _write_xlsx(self, rows: list[dict], name: str = "linked-source.xlsx") -> Path:
        path = self.root / name
        workbook = Workbook()
        ignored = workbook.active
        ignored.title = "Ignored"
        ignored.append(CSV_COLUMNS)
        ignored.append(list(self._row("wrong-sheet", "wrong-sheet").values()))
        selected = workbook.create_sheet("Vocabulary")
        columns = list(dict.fromkeys(CSV_COLUMNS + [key for row in rows for key in row]))
        selected.append(columns)
        for row in rows:
            selected.append([row.get(column, "") for column in columns])
        workbook.save(path)
        return path

    def _counts(self) -> dict[str, int]:
        names = (
            "entries",
            "entry_collections",
            "cards",
            "card_revisions",
            "card_revision_entries",
            "collection_source_links",
        )
        with db.get_connection() as conn:
            return {
                name: int(conn.execute(f"SELECT COUNT(*) FROM {name}").fetchone()[0])
                for name in names
            }

    def _revision_count(self) -> int:
        with db.get_connection() as conn:
            return int(
                conn.execute(
                    """
                    SELECT COUNT(*)
                    FROM card_revisions
                    JOIN cards ON cards.id = card_revisions.card_id
                    WHERE cards.collection_id = ?
                    """,
                    (self.collection_id,),
                ).fetchone()[0]
            )

    def _terms(self) -> list[str]:
        return [str(row["term"]) for row in get_entries_in_collection(self.collection_id)]

    def test_fresh_schema_contains_small_link_table_and_is_idempotent(self) -> None:
        db.init_db()
        db.init_db()
        with db.get_connection() as conn:
            columns = [row["name"] for row in conn.execute("PRAGMA table_info(collection_source_links)")]
            foreign_keys = conn.execute("PRAGMA foreign_key_list(collection_source_links)").fetchall()
            self.assertEqual(get_schema_version(conn), CURRENT_SCHEMA_VERSION)
            self.assertEqual(get_metadata(conn, "app_data_version"), APP_DATA_VERSION)
            self.assertEqual(conn.execute("SELECT COUNT(*) FROM collection_source_links").fetchone()[0], 0)
        self.assertEqual(
            columns,
            [
                "collection_id",
                "source_path",
                "source_type",
                "import_mode",
                "sheet_name",
                "linked_at",
                "last_refreshed_at",
            ],
        )
        self.assertEqual(
            [(row["table"], row["from"], row["on_delete"]) for row in foreign_keys],
            [("collections", "collection_id", "CASCADE")],
        )

    def test_pre_m13_data_migrates_and_schema_converges(self) -> None:
        with db.get_connection() as fresh:
            fresh_schema = [tuple(row) for row in fresh.execute("PRAGMA table_info(collection_source_links)")]
        migrated_path = self.root / "pre-m13.sqlite3"
        db.DB_PATH = migrated_path
        db.init_db()
        with db.get_connection() as conn:
            conn.execute("DROP TABLE collection_source_links")
            set_schema_version(conn, QUIZ_LOG_HISTORY_SCHEMA_VERSION)
            set_metadata(conn, "app_data_version", "11.3")
            conn.execute(
                "INSERT INTO collections (name, description, card_size, created_at, updated_at) VALUES ('Preserved', '', 8, 'now', 'now')"
            )
            applied = run_migrations(conn)
            migrated_schema = [tuple(row) for row in conn.execute("PRAGMA table_info(collection_source_links)")]
            preserved = conn.execute("SELECT name FROM collections WHERE name = 'Preserved'").fetchone()
            self.assertEqual(get_schema_version(conn), CURRENT_SCHEMA_VERSION)
            self.assertEqual(get_metadata(conn, "app_data_version"), APP_DATA_VERSION)
        self.assertEqual(applied, ["m13_linked_append_source", "m15.1_template_speech_semantics"])
        self.assertEqual(migrated_schema, fresh_schema)
        self.assertIsNotNone(preserved)

    def test_migration_failure_rolls_back_table_and_metadata(self) -> None:
        with db.get_connection() as conn:
            conn.execute("DROP TABLE collection_source_links")
            set_schema_version(conn, QUIZ_LOG_HISTORY_SCHEMA_VERSION)
            set_metadata(conn, "app_data_version", "11.3")

            def failing_migration(connection: sqlite3.Connection) -> None:
                connection.execute("CREATE TABLE collection_source_links (collection_id INTEGER)")
                set_metadata(connection, "app_data_version", "13.0")
                raise RuntimeError("synthetic migration failure")

            migration = next(
                item for item in MIGRATIONS if item["name"] == "m13_linked_append_source"
            )
            original = migration["function"]
            migration["function"] = failing_migration
            try:
                with self.assertRaises(RuntimeError):
                    run_migrations(conn)
            finally:
                migration["function"] = original
            table = conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = 'collection_source_links'"
            ).fetchone()
            self.assertIsNone(table)
            self.assertEqual(get_schema_version(conn), QUIZ_LOG_HISTORY_SCHEMA_VERSION)
            self.assertEqual(get_metadata(conn, "app_data_version"), "11.3")

    def test_csv_initial_preview_is_read_only_and_classified(self) -> None:
        source = self._write_csv(
            [self._row("new", "valid"), self._row("invalid", "", status="new")]
        )
        before = self._counts()
        preview = preview_collection_source_link(
            self.collection_id, source, "general_entry"
        )
        self.assertTrue(preview["can_confirm"])
        self.assertEqual(preview["summary"], {
            "total_rows": 2,
            "new_valid_count": 1,
            "invalid_count": 1,
            "duplicate_count": 0,
        })
        self.assertEqual(self._counts(), before)

    def test_xlsx_preview_is_read_only_and_persists_selected_sheet(self) -> None:
        source = self._write_xlsx([self._row("selected", "sheet")])
        before = self._counts()
        preview = preview_collection_source_link(
            self.collection_id, source, "general_entry", sheet_name="Vocabulary"
        )
        self.assertTrue(preview["ok"])
        self.assertEqual(preview["new_valid_rows"][0]["data"]["term"], "selected")
        self.assertEqual(self._counts(), before)
        confirmed = confirm_collection_source_link(
            self.collection_id, source, "general_entry", sheet_name="Vocabulary"
        )
        self.assertTrue(confirmed["success"])
        self.assertEqual(get_collection_source_link(self.collection_id)["sheet_name"], "Vocabulary")
        self.assertEqual(preview_linked_source_refresh(self.collection_id)["summary"]["new_valid_count"], 0)

    def test_missing_unsupported_and_missing_worksheet_are_controlled(self) -> None:
        missing = preview_collection_source_link(
            self.collection_id, self.root / "missing.csv", "general_entry"
        )
        unsupported = preview_collection_source_link(
            self.collection_id, self.root / "unsupported.txt", "general_entry"
        )
        source = self._write_xlsx([self._row("term", "meaning")])
        worksheet = preview_collection_source_link(
            self.collection_id, source, "general_entry", sheet_name="Missing"
        )
        self.assertEqual(missing["errors"], ["Linked source is unavailable."])
        self.assertEqual(unsupported["errors"], ["Linked source format is no longer supported."])
        self.assertEqual(worksheet["errors"], ["Worksheet is unavailable."])
        self.assertEqual(self._counts()["entries"], 0)

    def test_initial_confirm_imports_new_only_and_persists_link_after_confirmation(self) -> None:
        existing_source = self._write_csv([self._row("existing", "meaning")], "seed.csv")
        seed = confirm_collection_source_link(
            self.collection_id, existing_source, "general_entry"
        )
        self.assertEqual(seed["imported_count"], 1)
        unlink_collection_source(self.collection_id)
        source = self._write_csv(
            [
                self._row("existing", "meaning"),
                self._row("new", "meaning"),
                self._row("invalid", ""),
            ]
        )
        preview = preview_collection_source_link(self.collection_id, source, "general_entry")
        self.assertEqual(preview["summary"]["duplicate_count"], 1)
        self.assertEqual(preview["summary"]["new_valid_count"], 1)
        self.assertIsNone(get_collection_source_link(self.collection_id))
        result = confirm_collection_source_link(self.collection_id, source, "general_entry")
        self.assertTrue(result["success"])
        self.assertEqual(result["imported_count"], 1)
        self.assertEqual(self._terms(), ["existing", "new"])
        self.assertNotIn("import_anyway", repr(result))

    def test_zero_new_valid_can_still_establish_link(self) -> None:
        source = self._write_csv([self._row("existing", "meaning")])
        import_general_entry_rows(
            [{"row_number": 2, "errors": [], "data": self._row("existing", "meaning")}],
            target_collection_id=self.collection_id,
        )
        result = confirm_collection_source_link(self.collection_id, source, "general_entry")
        self.assertTrue(result["success"])
        self.assertEqual(result["imported_count"], 0)
        self.assertIsNotNone(get_collection_source_link(self.collection_id))

    def test_second_link_is_rejected_without_replacement(self) -> None:
        first = self._write_csv([self._row("first", "meaning")], "first.csv")
        second = self._write_csv([self._row("second", "meaning")], "second.csv")
        confirm_collection_source_link(self.collection_id, first, "general_entry")
        preview = preview_collection_source_link(self.collection_id, second, "general_entry")
        confirm = confirm_collection_source_link(self.collection_id, second, "general_entry")
        self.assertFalse(preview["can_confirm"])
        self.assertFalse(confirm["success"])
        self.assertEqual(Path(get_collection_source_link(self.collection_id)["source_path"]), first.resolve())
        self.assertNotIn("second", self._terms())

    def test_unchanged_and_appended_refresh_are_idempotent(self) -> None:
        source = self._write_csv([self._row("first", "meaning")])
        confirm_collection_source_link(self.collection_id, source, "general_entry")
        revisions = self._revision_count()
        unchanged = confirm_linked_source_refresh(self.collection_id)
        self.assertEqual(unchanged["imported_count"], 0)
        self.assertEqual(self._revision_count(), revisions)
        self._write_csv([self._row("first", "meaning"), self._row("second", "meaning")])
        preview = preview_linked_source_refresh(self.collection_id)
        self.assertEqual(preview["summary"]["new_valid_count"], 1)
        appended = confirm_linked_source_refresh(self.collection_id)
        self.assertEqual(appended["imported_count"], 1)
        self.assertEqual(self._terms(), ["first", "second"])
        self.assertEqual(confirm_linked_source_refresh(self.collection_id)["imported_count"], 0)

    def test_source_deletion_and_reorder_do_not_delete_or_reorder_app_content(self) -> None:
        rows = [self._row("one", "1"), self._row("two", "2"), self._row("three", "3")]
        source = self._write_csv(rows)
        confirm_collection_source_link(self.collection_id, source, "general_entry")
        before_terms = self._terms()
        before_revisions = self._revision_count()
        self._write_csv([rows[2], rows[0]])
        result = confirm_linked_source_refresh(self.collection_id)
        self.assertEqual(result["imported_count"], 0)
        self.assertEqual(self._terms(), before_terms)
        self.assertEqual(self._revision_count(), before_revisions)

    def test_source_edit_never_overwrites_and_may_append_new_content(self) -> None:
        source = self._write_csv([self._row("original", "meaning")])
        confirm_collection_source_link(self.collection_id, source, "general_entry")
        original_id = int(get_entries_in_collection(self.collection_id)[0]["id"])
        self._write_csv([self._row("edited", "meaning")])
        preview = preview_linked_source_refresh(self.collection_id)
        self.assertEqual(preview["summary"]["new_valid_count"], 1)
        confirm_linked_source_refresh(self.collection_id)
        self.assertEqual(get_entry_by_id(original_id)["term"], "original")
        self.assertEqual(self._terms(), ["original", "edited"])

    def test_app_edit_and_delete_never_write_source_and_deleted_row_can_reappear(self) -> None:
        source = self._write_csv([self._row("source-term", "source-meaning")])
        original_bytes = source.read_bytes()
        confirm_collection_source_link(self.collection_id, source, "general_entry")
        entry_id = int(get_entries_in_collection(self.collection_id)[0]["id"])
        update_entry(
            entry_id,
            "English",
            "English",
            "word",
            "app-edited",
            "source-meaning",
        )
        self.assertEqual(source.read_bytes(), original_bytes)
        delete_entry(entry_id)
        self.assertEqual(source.read_bytes(), original_bytes)
        preview = preview_linked_source_refresh(self.collection_id)
        self.assertEqual(preview["summary"]["new_valid_count"], 1)
        confirm_linked_source_refresh(self.collection_id)
        self.assertEqual(self._terms(), ["source-term"])

    def test_unavailable_refresh_preserves_content_link_history_and_timestamp(self) -> None:
        source = self._write_csv([self._row("stable", "meaning")])
        confirm_collection_source_link(self.collection_id, source, "general_entry")
        confirm_linked_source_refresh(self.collection_id)
        link_before = get_collection_source_link(self.collection_id)
        counts_before = self._counts()
        source.unlink()
        preview = preview_linked_source_refresh(self.collection_id)
        result = confirm_linked_source_refresh(self.collection_id)
        self.assertFalse(preview["ok"])
        self.assertFalse(result["success"])
        self.assertEqual(get_collection_source_link(self.collection_id), link_before)
        self.assertEqual(self._counts(), counts_before)
        self.assertEqual(self._terms(), ["stable"])

    def test_unlink_and_collection_delete_remove_metadata_only(self) -> None:
        source = self._write_csv([self._row("stable", "meaning")])
        confirm_collection_source_link(self.collection_id, source, "general_entry")
        entry_id = int(get_entries_in_collection(self.collection_id)[0]["id"])
        result = unlink_collection_source(self.collection_id)
        repeated = unlink_collection_source(self.collection_id)
        self.assertTrue(result["unlinked"])
        self.assertFalse(repeated["unlinked"])
        self.assertIsNotNone(get_entry_by_id(entry_id))
        self.assertEqual(self._terms(), ["stable"])

        confirm_collection_source_link(self.collection_id, source, "general_entry")
        delete_collection(self.collection_id)
        with db.get_connection() as conn:
            self.assertEqual(conn.execute("SELECT COUNT(*) FROM collection_source_links").fetchone()[0], 0)
        self.assertIsNotNone(get_entry_by_id(entry_id))

    def test_initial_confirm_failure_rolls_back_everything(self) -> None:
        source = self._write_csv([self._row("rollback", "meaning")])
        before = self._counts()

        def partial_failure(rows, **kwargs):
            connection = kwargs["connection"]
            connection.execute(
                """
                INSERT INTO entries (
                    language, explanation_language, entry_type, term, meaning,
                    status, created_at, updated_at
                ) VALUES ('English', 'English', 'word', 'partial', 'partial', 'new', 'now', 'now')
                """
            )
            raise RuntimeError("synthetic write failure")

        with patch("src.linked_sources.import_general_entry_rows", side_effect=partial_failure):
            result = confirm_collection_source_link(self.collection_id, source, "general_entry")
        self.assertFalse(result["success"])
        self.assertEqual(self._counts(), before)
        self.assertIsNone(get_collection_source_link(self.collection_id))

    def test_refresh_failure_rolls_back_changes_and_timestamp(self) -> None:
        source = self._write_csv([self._row("stable", "meaning")])
        confirm_collection_source_link(self.collection_id, source, "general_entry")
        confirm_linked_source_refresh(self.collection_id)
        link_before = get_collection_source_link(self.collection_id)
        counts_before = self._counts()
        self._write_csv([self._row("stable", "meaning"), self._row("partial", "meaning")])

        def partial_failure(rows, **kwargs):
            connection = kwargs["connection"]
            cursor = connection.execute(
                """
                INSERT INTO entries (
                    language, explanation_language, entry_type, term, meaning,
                    status, created_at, updated_at
                ) VALUES ('English', 'English', 'word', 'partial', 'meaning', 'new', 'now', 'now')
                """
            )
            connection.execute(
                "INSERT INTO entry_collections (entry_id, collection_id, position, added_at) VALUES (?, ?, 2, 'now')",
                (int(cursor.lastrowid), self.collection_id),
            )
            raise RuntimeError("synthetic refresh failure")

        with patch("src.linked_sources.import_general_entry_rows", side_effect=partial_failure):
            result = confirm_linked_source_refresh(self.collection_id)
        self.assertFalse(result["success"])
        self.assertEqual(self._counts(), counts_before)
        self.assertEqual(get_collection_source_link(self.collection_id), link_before)
        self.assertEqual(self._terms(), ["stable"])

    def test_caller_owned_import_reconciles_history_without_committing(self) -> None:
        connection = db.get_connection()
        try:
            before = self._revision_count()
            result = import_general_entry_rows(
                [{"row_number": 2, "errors": [], "data": self._row("caller", "owned")}],
                target_collection_id=self.collection_id,
                connection=connection,
            )
            self.assertEqual(result["imported_count"], 1)
            in_transaction_revisions = int(
                connection.execute(
                    """
                    SELECT COUNT(*) FROM card_revisions
                    JOIN cards ON cards.id = card_revisions.card_id
                    WHERE cards.collection_id = ?
                    """,
                    (self.collection_id,),
                ).fetchone()[0]
            )
            self.assertEqual(in_transaction_revisions, before + 1)
            connection.rollback()
        finally:
            connection.close()
        self.assertEqual(self._revision_count(), before)
        self.assertEqual(self._terms(), [])

    def test_append_creates_only_required_card_revisions(self) -> None:
        source = self._write_csv([self._row("one", "1")])
        confirm_collection_source_link(self.collection_id, source, "general_entry")
        first_revisions = self._revision_count()
        self.assertEqual(first_revisions, 1)
        self._write_csv([self._row("one", "1"), self._row("two", "2")])
        confirm_linked_source_refresh(self.collection_id)
        self.assertEqual(self._revision_count(), first_revisions + 1)
        self._write_csv([
            self._row("one", "1"),
            self._row("two", "2"),
            self._row("three", "3"),
        ])
        confirm_linked_source_refresh(self.collection_id)
        self.assertEqual(self._revision_count(), first_revisions + 2)

    def test_template_aware_mode_resolves_current_template_and_missing_template_is_invalid(self) -> None:
        template_id = create_entry_template(
            "Synthetic Portable",
            "Synthetic",
            "English",
            "synthetic_portable",
        )
        create_template_field(template_id, "term", "Term", "text", True, 0)
        create_template_field(template_id, "meaning", "Meaning", "text", True, 0)
        rows = [{
            "language": "English",
            "template_name": "Synthetic Portable",
            "template_type": "synthetic_portable",
            "field:term": "portable",
            "field:meaning": "meaning",
        }]
        source = self._write_csv(rows, "template-aware.csv")
        result = confirm_collection_source_link(
            self.collection_id, source, "template_aware"
        )
        self.assertTrue(result["success"])
        self.assertEqual(self._terms(), ["portable"])
        with db.get_connection() as conn:
            conn.execute("DELETE FROM entry_templates WHERE id = ?", (template_id,))
        preview = preview_linked_source_refresh(self.collection_id)
        self.assertTrue(preview["ok"])
        self.assertEqual(preview["summary"]["invalid_count"], 1)
        self.assertEqual(preview["summary"]["new_valid_count"], 0)
        self.assertEqual(confirm_linked_source_refresh(self.collection_id)["imported_count"], 0)

    def test_database_and_xlsx_backups_include_link_metadata(self) -> None:
        source = self._write_csv([self._row("backup", "meaning")])
        confirm_collection_source_link(self.collection_id, source, "general_entry")
        self.assertIn("collection_source_links", BACKUP_TABLES)
        workbook_bytes = build_full_backup_workbook_bytes()
        workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
        self.assertIn("collection_source_links", workbook.sheetnames)
        link_rows = list(workbook["collection_source_links"].iter_rows(values_only=True))
        self.assertEqual(link_rows[0][0], "collection_id")
        self.assertEqual(link_rows[1][0], self.collection_id)
        backup_path = self.root / "backup.sqlite3"
        backup_path.write_bytes(get_database_file_bytes())
        restored = sqlite3.connect(backup_path)
        try:
            restored.row_factory = sqlite3.Row
            link = restored.execute(
                "SELECT collection_id, source_type, import_mode FROM collection_source_links"
            ).fetchone()
            self.assertEqual(dict(link), {
                "collection_id": self.collection_id,
                "source_type": "csv",
                "import_mode": "general_entry",
            })
            self.assertEqual(restored.execute("PRAGMA integrity_check").fetchone()[0], "ok")
        finally:
            restored.close()


if __name__ == "__main__":
    unittest.main()
