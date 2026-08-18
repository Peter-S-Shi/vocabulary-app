from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

try:
    from PySide6.QtWidgets import QApplication

    PYSIDE6_AVAILABLE = True
except ImportError:  # pragma: no cover - exercised only when PySide6 is absent
    PYSIDE6_AVAILABLE = False

from src import db
from src.backup import build_full_backup_workbook_bytes, preview_backup_workbook
from src.collections import create_collection, get_entries_in_collection
from src.linked_sources import get_collection_source_link

"""
M19 Phase B/C adversarial boundary hardening: backup/restore-preview and
Linked Source under deliberately hostile, non-happy-path input that the
M18 suites did not cover -- truncated/corrupted workbook bytes, tampered
backup metadata, and the linked file changing (vanishing, corrupting, or
being rewritten) between the user's Preview and their Confirm.

These lock in the safety semantics the core already promises: restore
stays preview-only and never mutates the active database on bad input;
linked-source Confirm re-validates the file inside its own transaction
and fails closed (rollback, link metadata and Entries untouched) when
the file is gone or unreadable; and because M13 v1 stores no source-row
identity, Confirm appends what the file contains at confirmation time --
asserted here so any future silent drift of that semantic fails a test.
"""

if PYSIDE6_AVAILABLE:
    from src.ui_desktop.controllers.data_tools_controller import DataToolsController
    from src.ui_desktop.controllers.linked_source_controller import LinkedSourceController

    def _qt_app() -> QApplication:
        app = QApplication.instance()
        if app is None:
            app = QApplication([])
        return app


class _SyntheticDatabaseTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory(ignore_cleanup_errors=True)
        self.root = Path(self.temp_dir.name)
        self.original_db_path = db.DB_PATH
        db.DB_PATH = self.root / "m19_adversarial.sqlite3"
        db.init_db()

    def tearDown(self) -> None:
        db.DB_PATH = self.original_db_path
        self.temp_dir.cleanup()

    def _table_counts(self) -> dict[str, int]:
        conn = db.get_connection()
        try:
            tables = [
                row[0]
                for row in conn.execute(
                    "select name from sqlite_master where type='table' and name != 'sqlite_sequence'"
                )
            ]
            return {table: conn.execute(f"select count(*) from {table}").fetchone()[0] for table in tables}
        finally:
            conn.close()

    def _write_source_csv(self, name: str, rows: str) -> str:
        path = self.root / name
        path.write_text("language,term,meaning\n" + rows, encoding="utf-8")
        return str(path)


class BackupPreviewHostileBytesTests(_SyntheticDatabaseTestCase):
    def test_truncated_workbook_bytes_report_invalid_without_mutation(self) -> None:
        counts_before = self._table_counts()
        real = build_full_backup_workbook_bytes()
        truncated = real[: len(real) // 3]

        preview = preview_backup_workbook(truncated)

        self.assertFalse(preview["valid_backup"])
        self.assertTrue(preview["errors"])
        self.assertEqual(self._table_counts(), counts_before)

    def test_garbage_bytes_report_invalid_not_a_crash(self) -> None:
        preview = preview_backup_workbook(b"\x00\x01\x02 not a zip archive \xff" * 100)
        self.assertFalse(preview["valid_backup"])
        self.assertTrue(preview["errors"])

    def test_tampered_metadata_app_name_surfaces_a_warning(self) -> None:
        """Existing core contract, locked in: a foreign app_name is a
        WARNING, not an error -- restore is preview-only, so the honest
        response is to show the mismatch, not to hard-reject a file the
        user may legitimately want to inspect. An unsupported
        backup_format_version remains a hard error (covered by
        validate_backup_metadata)."""
        from io import BytesIO

        from openpyxl import load_workbook

        real = build_full_backup_workbook_bytes()
        workbook = load_workbook(BytesIO(real))
        metadata_sheet = workbook["backup_metadata"]
        for row in metadata_sheet.iter_rows(min_row=2):
            if row[0].value == "app_name":
                row[1].value = "Some Other Application"
        buffer = BytesIO()
        workbook.save(buffer)

        preview = preview_backup_workbook(buffer.getvalue())

        self.assertTrue(any("app name" in warning.lower() for warning in preview["warnings"]))

    def test_unsupported_format_version_is_a_hard_error(self) -> None:
        from io import BytesIO

        from openpyxl import load_workbook

        real = build_full_backup_workbook_bytes()
        workbook = load_workbook(BytesIO(real))
        metadata_sheet = workbook["backup_metadata"]
        for row in metadata_sheet.iter_rows(min_row=2):
            if row[0].value == "backup_format_version":
                row[1].value = "999.0"
        buffer = BytesIO()
        workbook.save(buffer)

        preview = preview_backup_workbook(buffer.getvalue())

        self.assertFalse(preview["valid_backup"])
        self.assertTrue(any("version" in error.lower() for error in preview["errors"]))


@unittest.skipUnless(PYSIDE6_AVAILABLE, "PySide6 is not installed; see requirements-desktop.txt.")
class RestorePreviewControllerHostileBytesTests(_SyntheticDatabaseTestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.app = _qt_app()

    def test_controller_surfaces_hostile_bytes_as_controlled_state(self) -> None:
        counts_before = self._table_counts()
        controller = DataToolsController()
        controller.load_restore_preview_file(b"PK\x03\x04 corrupted beyond parsing", "broken.xlsx")
        controller.run_restore_preview()

        self.assertIsNotNone(controller.restore_preview_result)
        self.assertFalse(controller.restore_preview_result["valid_backup"])
        self.assertEqual(self._table_counts(), counts_before)


@unittest.skipUnless(PYSIDE6_AVAILABLE, "PySide6 is not installed; see requirements-desktop.txt.")
class LinkedSourceFileMutatedBetweenPreviewAndConfirmTests(_SyntheticDatabaseTestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.app = _qt_app()

    def _linked_controller(self, collection_id: int, source_path: str) -> "LinkedSourceController":
        controller = LinkedSourceController()
        controller.open_for_collection(collection_id, "Fruits")
        controller.stage_source_path(source_path)
        controller.run_preview()
        controller.confirm()
        return controller

    def test_file_deleted_between_preview_and_confirm_fails_closed(self) -> None:
        collection_id = create_collection("Fruits", "", card_size=8)
        source_path = self._write_source_csv("fruits.csv", "French,pomme,apple\n")
        controller = self._linked_controller(collection_id, source_path)
        link_before = get_collection_source_link(collection_id)

        Path(source_path).write_text("language,term,meaning\nFrench,poire,pear\n", encoding="utf-8")
        controller.open_for_collection(collection_id, "Fruits")
        controller.run_preview()
        self.assertTrue(controller.can_confirm())

        os.remove(source_path)
        result = controller.confirm()

        self.assertFalse(result["success"])
        self.assertEqual(get_collection_source_link(collection_id), link_before)
        self.assertEqual(len(get_entries_in_collection(collection_id)), 1)

    def test_file_corrupted_between_preview_and_confirm_fails_closed(self) -> None:
        collection_id = create_collection("Fruits", "", card_size=8)
        source_path = self._write_source_csv("fruits.csv", "French,pomme,apple\n")
        controller = self._linked_controller(collection_id, source_path)

        with open(source_path, "a", encoding="utf-8") as handle:
            handle.write("French,poire,pear\n")
        controller.open_for_collection(collection_id, "Fruits")
        controller.run_preview()
        self.assertTrue(controller.can_confirm())

        # Replace the CSV with bytes that cannot be parsed as text rows.
        Path(source_path).write_bytes(b"\x00\xff\x00\xff garbage")
        result = controller.confirm()

        self.assertFalse(result["success"])
        self.assertEqual(len(get_entries_in_collection(collection_id)), 1)

    def test_confirm_applies_confirmation_time_file_content(self) -> None:
        """M13 v1 semantic, locked in: Confirm re-scans the file inside
        its own transaction, so it appends what the file contains at
        confirmation time (still New-Valid-only, still append-only). If
        this ever silently changes -- in either direction -- a deliberate
        decision should be recorded, not an accident."""
        collection_id = create_collection("Fruits", "", card_size=8)
        source_path = self._write_source_csv("fruits.csv", "French,pomme,apple\n")
        controller = self._linked_controller(collection_id, source_path)

        with open(source_path, "a", encoding="utf-8") as handle:
            handle.write("French,poire,pear\n")
        controller.open_for_collection(collection_id, "Fruits")
        controller.run_preview()
        self.assertEqual(controller.preview["summary"]["new_valid_count"], 1)

        # The file gains one more new row after the preview was shown.
        with open(source_path, "a", encoding="utf-8") as handle:
            handle.write("French,peche,peach\n")
        result = controller.confirm()

        self.assertTrue(result["success"])
        terms = {entry["term"] for entry in get_entries_in_collection(collection_id)}
        self.assertEqual(terms, {"pomme", "poire", "peche"})


if __name__ == "__main__":  # pragma: no cover
    unittest.main()
