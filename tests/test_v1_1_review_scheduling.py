from __future__ import annotations

from pathlib import Path
from datetime import date
from io import BytesIO
import os
import tempfile
import unittest
from unittest.mock import patch

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PySide6.QtCore import QDate
from PySide6.QtWidgets import QApplication, QLabel, QMessageBox, QPushButton, QWidget
from openpyxl import load_workbook

from src import db, quiz
from src.app_config import get_default_db_path
from src.backup import BACKUP_TABLES, build_full_backup_workbook_bytes, preview_backup_workbook
from src.card_history import reconcile_collection_card_history
from src.learning_workflow import get_card_learning_history
from src.migrations import (
    CURRENT_SCHEMA_VERSION,
    REVIEW_SCHEDULE_SCHEMA_VERSION,
    SPEECH_SEMANTICS_SCHEMA_VERSION,
    get_compatibility_status,
    run_migrations,
    set_metadata,
    set_schema_version,
)
from src.review import get_card_review_state
from src.review_schedule import (
    clear_card_schedule,
    get_card_schedule,
    list_card_schedules,
    list_actionable_schedules,
    schedule_card_after_days,
    set_card_next_review,
)
from src.ui_desktop.controllers.quiz_controller import QuizController
from src.ui_desktop.controllers.today_controller import TodayController
from src.ui_desktop.controllers.review_calendar_controller import ReviewCalendarController
from src.ui_desktop.views.review_calendar_view import ReviewCalendarView
from src.ui_desktop.views.quiz_view import QuizView
from src.ui_desktop.views.today_view import TodayView


class ReviewScheduleTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory(ignore_cleanup_errors=True)
        self.original_db_path = db.DB_PATH
        db.DB_PATH = Path(self.temp_dir.name) / "review-schedule.sqlite3"
        db.init_db()

    def tearDown(self) -> None:
        db.DB_PATH = self.original_db_path
        self.temp_dir.cleanup()

    def _card_id(self, *, name: str = "Synthetic Review") -> int:
        if db.DB_PATH.expanduser().resolve() == get_default_db_path().resolve():
            raise AssertionError(
                "Synthetic review-scheduling fixtures require an isolated test database."
            )
        now = "2026-08-26T12:00:00+00:00"
        with db.get_connection() as conn:
            collection_id = int(
                conn.execute(
                    """
                    INSERT INTO collections (name, description, card_size, created_at, updated_at)
                    VALUES (?, '', 1, ?, ?)
                    """,
                    (name, now, now),
                ).lastrowid
            )
            entry_id = int(
                conn.execute(
                    """
                    INSERT INTO entries (
                        language, explanation_language, entry_type, term, meaning,
                        example, notes, tags, source, status, created_at, updated_at
                    ) VALUES ('English', 'English', 'word', 'term', 'meaning',
                              '', '', '', '', 'new', ?, ?)
                    """,
                    (now, now),
                ).lastrowid
            )
            conn.execute(
                """
                INSERT INTO entry_collections (entry_id, collection_id, position, added_at)
                VALUES (?, ?, 1, ?)
                """,
                (entry_id, collection_id, now),
            )
            reconcile_collection_card_history(
                conn,
                collection_id,
                change_reason="synthetic_review_schedule",
            )
            return int(
                conn.execute(
                    """
                    SELECT id FROM cards
                    WHERE collection_id = ? AND card_number = 1 AND is_active = 1
                    """,
                    (collection_id,),
                ).fetchone()[0]
            )

    def test_synthetic_fixture_rejects_production_database_path(self) -> None:
        production_db_path = get_default_db_path()
        with patch.object(db, "DB_PATH", production_db_path), \
                patch.object(db, "get_connection") as get_connection:
            with self.assertRaisesRegex(AssertionError, "isolated test database"):
                self._card_id(name="Synthetic Must Not Reach Production")

        get_connection.assert_not_called()


class StableCardScheduleTests(ReviewScheduleTestCase):
    def test_structured_backup_includes_review_schedules_and_consistent_metadata(self) -> None:
        card_id = self._card_id(name="Synthetic Schedule Backup")
        set_card_next_review(card_id, "2026-09-04", today="2026-08-26")

        backup_bytes = build_full_backup_workbook_bytes()
        workbook = load_workbook(BytesIO(backup_bytes), read_only=True, data_only=True)
        schedule_rows = list(
            workbook["card_review_schedules"].iter_rows(values_only=True)
        )
        metadata = dict(
            list(workbook["backup_metadata"].iter_rows(values_only=True))[1:]
        )
        preview = preview_backup_workbook(backup_bytes)
        schedule_preview = next(
            sheet
            for sheet in preview["sheets"]
            if sheet["sheet_name"] == "card_review_schedules"
        )

        headers = schedule_rows[0]
        exported = dict(zip(headers, schedule_rows[1], strict=True))
        self.assertEqual(exported["card_id"], card_id)
        self.assertEqual(exported["next_due_at"], "2026-09-04")
        self.assertEqual(metadata["table_count"], len(BACKUP_TABLES))
        self.assertEqual(schedule_preview["row_count"], 1)

    def test_clear_schedule_is_idempotent_and_preserves_quiz_history(self) -> None:
        card_id = self._card_id(name="Synthetic Clear Schedule")
        with db.get_connection() as conn:
            card = conn.execute(
                "SELECT collection_id, card_number FROM cards WHERE id = ?",
                (card_id,),
            ).fetchone()
            entry_id = int(
                conn.execute(
                    """
                    SELECT membership.entry_id
                    FROM card_revisions AS revision
                    JOIN card_revision_entries AS membership
                      ON membership.revision_id = revision.id
                    WHERE revision.card_id = ?
                    ORDER BY revision.revision_number DESC
                    LIMIT 1
                    """,
                    (card_id,),
                ).fetchone()[0]
            )
        session_id = quiz.create_quiz_session(
            int(card["collection_id"]),
            int(card["card_number"]),
            "term_to_meaning",
            1,
        )
        quiz.record_quiz_answer(
            session_id, entry_id, "term", "meaning", "meaning", True
        )
        quiz.complete_quiz_session(session_id)
        with db.get_connection() as conn:
            history_before = get_card_learning_history(
                conn, int(card["collection_id"]), int(card["card_number"])
            )

        set_card_next_review(card_id, "2026-08-26", today="2026-08-26")
        cleared = clear_card_schedule(card_id, today="2026-08-26")
        cleared_again = clear_card_schedule(card_id, today="2026-08-26")

        self.assertEqual(cleared["state"], "unscheduled")
        self.assertEqual(cleared_again["state"], "unscheduled")
        self.assertEqual(list_actionable_schedules(today="2026-08-26"), [])
        with db.get_connection() as conn:
            history_after = get_card_learning_history(
                conn, int(card["collection_id"]), int(card["card_number"])
            )
        self.assertEqual(history_after, history_before)

    def test_direct_schedule_calculation_uses_explicit_base_date(self) -> None:
        card_id = self._card_id(name="Synthetic Direct Schedule")

        scheduled = schedule_card_after_days(
            card_id,
            7,
            today="2026-08-26",
        )

        self.assertEqual(scheduled["next_due_at"], "2026-09-02")
        self.assertEqual(scheduled["state"], "upcoming")

    def test_retired_schedule_does_not_transfer_to_reused_card_number(self) -> None:
        old_card_id = self._card_id(name="Synthetic Reused Number")
        set_card_next_review(old_card_id, "2026-08-26", today="2026-08-26")
        now = "2026-08-26T13:00:00+00:00"
        with db.get_connection() as conn:
            collection_id = int(
                conn.execute(
                    "SELECT collection_id FROM cards WHERE id = ?",
                    (old_card_id,),
                ).fetchone()[0]
            )
            conn.execute(
                "DELETE FROM entry_collections WHERE collection_id = ?",
                (collection_id,),
            )
            reconcile_collection_card_history(
                conn,
                collection_id,
                change_reason="synthetic_retire",
            )
            replacement_entry_id = int(
                conn.execute(
                    """
                    INSERT INTO entries (
                        language, explanation_language, entry_type, term, meaning,
                        example, notes, tags, source, status, created_at, updated_at
                    ) VALUES ('English', 'English', 'word', 'replacement', 'new meaning',
                              '', '', '', '', 'new', ?, ?)
                    """,
                    (now, now),
                ).lastrowid
            )
            conn.execute(
                """
                INSERT INTO entry_collections (entry_id, collection_id, position, added_at)
                VALUES (?, ?, 1, ?)
                """,
                (replacement_entry_id, collection_id, now),
            )
            reconcile_collection_card_history(
                conn,
                collection_id,
                change_reason="synthetic_replacement",
            )
            replacement_card_id = int(
                conn.execute(
                    """
                    SELECT id FROM cards
                    WHERE collection_id = ? AND card_number = 1 AND is_active = 1
                    """,
                    (collection_id,),
                ).fetchone()[0]
            )

        self.assertNotEqual(replacement_card_id, old_card_id)
        self.assertEqual(get_card_schedule(old_card_id, today="2026-08-26")["state"], "retired")
        self.assertEqual(get_card_schedule(replacement_card_id, today="2026-08-26")["state"], "unscheduled")
        self.assertEqual(list_actionable_schedules(today="2026-08-26"), [])

    def test_fresh_database_reports_v1_1_schedule_data_version(self) -> None:
        with db.get_connection() as conn:
            compatibility = get_compatibility_status(conn)

        self.assertEqual(CURRENT_SCHEMA_VERSION, REVIEW_SCHEDULE_SCHEMA_VERSION)
        self.assertEqual(compatibility["schema_version"], REVIEW_SCHEDULE_SCHEMA_VERSION)
        self.assertEqual(compatibility["app_data_version"], "21.1")

    def test_v1_0_style_migration_preserves_ambiguous_legacy_state(self) -> None:
        card_id = self._card_id(name="Synthetic Legacy Migration")
        now = "2026-08-26T12:00:00+00:00"
        with db.get_connection() as conn:
            card = conn.execute(
                "SELECT collection_id, card_number FROM cards WHERE id = ?",
                (card_id,),
            ).fetchone()
            conn.execute("DROP TABLE card_review_schedules")
            conn.execute(
                """
                INSERT INTO card_review_states (
                    collection_id, card_number, next_due_at, created_at, updated_at
                ) VALUES (?, ?, '2026-08-27', ?, ?)
                """,
                (int(card["collection_id"]), int(card["card_number"]), now, now),
            )
            set_schema_version(conn, SPEECH_SEMANTICS_SCHEMA_VERSION)
            set_metadata(conn, "app_data_version", "15.1")
            applied = run_migrations(conn)

        legacy = get_card_review_state(
            int(card["collection_id"]),
            int(card["card_number"]),
        )
        current = list_card_schedules(today="2026-08-26")
        self.assertEqual(applied, ["v1.1_stable_card_review_schedule"])
        self.assertEqual(legacy["next_due_at"], "2026-08-27")
        self.assertEqual(current[0]["card_id"], card_id)
        self.assertEqual(current[0]["state"], "unscheduled")

    def test_card_schedule_is_unscheduled_until_user_sets_next_review(self) -> None:
        card_id = self._card_id()

        self.assertEqual(get_card_schedule(card_id)["state"], "unscheduled")

        scheduled = set_card_next_review(
            card_id,
            "2026-08-28",
            today="2026-08-26",
        )

        self.assertEqual(scheduled["card_id"], card_id)
        self.assertEqual(scheduled["next_due_at"], "2026-08-28")
        self.assertEqual(scheduled["state"], "upcoming")
        self.assertEqual(
            get_card_schedule(card_id, today="2026-08-26"),
            scheduled,
        )

    def test_actionable_schedules_exclude_unscheduled_and_include_due_states(self) -> None:
        overdue_id = self._card_id(name="Synthetic Overdue")
        due_today_id = self._card_id(name="Synthetic Due Today")
        self._card_id(name="Synthetic Unscheduled")
        set_card_next_review(overdue_id, "2026-08-25", today="2026-08-26")
        set_card_next_review(due_today_id, "2026-08-26", today="2026-08-26")

        schedules = list_actionable_schedules(today="2026-08-26")

        self.assertEqual(
            [(row["card_id"], row["state"]) for row in schedules],
            [(overdue_id, "overdue"), (due_today_id, "due_today")],
        )


class QuizCompletionScheduleTests(ReviewScheduleTestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.app = QApplication.instance() or QApplication([])

    def test_completed_card_quiz_can_schedule_without_new_learning_event(self) -> None:
        card_id = self._card_id(name="Synthetic Quiz Schedule")
        with db.get_connection() as conn:
            card = conn.execute(
                "SELECT collection_id, card_number FROM cards WHERE id = ?",
                (card_id,),
            ).fetchone()
            entry_id = int(
                conn.execute(
                    """
                    SELECT membership.entry_id
                    FROM card_revisions AS revisions
                    JOIN card_revision_entries AS membership
                      ON membership.revision_id = revisions.id
                    WHERE revisions.card_id = ?
                    ORDER BY revisions.revision_number DESC
                    LIMIT 1
                    """,
                    (card_id,),
                ).fetchone()[0]
            )
        session_id = quiz.create_quiz_session(
            int(card["collection_id"]),
            int(card["card_number"]),
            "term_to_meaning",
            1,
        )
        quiz.record_quiz_answer(
            session_id,
            entry_id,
            "term",
            "meaning",
            "meaning",
            True,
        )
        completed = quiz.complete_quiz_session(session_id)
        controller = QuizController()
        controller.completed_session = completed
        with db.get_connection() as conn:
            before = get_card_learning_history(
                conn,
                int(card["collection_id"]),
                int(card["card_number"]),
            )

        scheduled = controller.schedule_next_review(
            "2026-09-02",
            today="2026-08-26",
        )

        self.assertEqual(scheduled["card_id"], card_id)
        self.assertEqual(scheduled["next_due_at"], "2026-09-02")
        with db.get_connection() as conn:
            after = get_card_learning_history(
                conn,
                int(card["collection_id"]),
                int(card["card_number"]),
            )
        self.assertEqual(after, before)

    def test_card_completion_view_offers_next_review_presets(self) -> None:
        card_id = self._card_id(name="Synthetic Quiz Completion UI")
        controller = QuizController(today_provider=lambda: date(2026, 8, 26))
        controller.completed_session = {
            "card_id": card_id,
            "card_number": 1,
            "total_items": 1,
            "correct_count": 1,
            "wrong_count": 0,
        }
        view = QuizView(controller)
        self.addCleanup(view.deleteLater)

        view._render()
        preset = view.findChild(QPushButton, "quiz-completion-schedule-7-days")
        self.assertIsNotNone(preset)
        preset.click()

        # Selection is staged, not immediately persisted to DB
        self.assertIsNone(controller.completion_schedule()["next_due_at"])

        # Clicking confirmation button persists the schedule
        save_button = view.findChild(QPushButton, "quiz-completion-schedule-save-button")
        self.assertIsNotNone(save_button)
        save_button.click()

        self.assertEqual(controller.completion_schedule()["next_due_at"], "2026-09-02")
        status_label = view.findChild(QLabel, "quiz-completion-schedule-status")
        self.assertIsNotNone(status_label)
        self.assertEqual(status_label.text(), "Scheduled · 2026-09-02")

    def test_non_card_completion_has_no_schedule_controls(self) -> None:
        controller = QuizController()
        controller.completed_session = {
            "card_id": None,
            "card_number": 0,
            "total_items": 1,
            "correct_count": 1,
            "wrong_count": 0,
        }
        view = QuizView(controller)
        self.addCleanup(view.deleteLater)

        view._render()

        self.assertIsNone(
            view.findChild(QPushButton, "quiz-completion-schedule-7-days")
        )


class TodayScheduleRoutingTests(ReviewScheduleTestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.app = QApplication.instance() or QApplication([])

    def test_due_schedule_is_an_actionable_card_quiz_queue_item(self) -> None:
        card_id = self._card_id(name="Synthetic Today Due")
        set_card_next_review(card_id, "2026-08-26", today="2026-08-26")
        controller = TodayController()
        view = TodayView(controller)
        self.addCleanup(view.deleteLater)
        launched = []
        view.quiz_launch_requested.connect(launched.append)

        overview = controller.refresh(today="2026-08-26")
        queue_item = controller.queue_items()[0]
        intent = controller.build_learning_action_intent(queue_item)
        queue_card = view.findChild(QWidget, "today-queue-card")
        title = queue_card.findChild(QLabel, "today-action-title")
        description = queue_card.findChild(QLabel, "today-action-subtitle")
        action = queue_card.findChild(QPushButton, "today-action-button")

        self.assertEqual(overview["due_schedules"][0]["card_id"], card_id)
        self.assertEqual(queue_item["recommendation_type"], "scheduled_review")
        self.assertTrue(title.text().strip())
        self.assertTrue(description.text().strip())
        self.assertTrue(action.isEnabled())
        self.assertEqual(intent.action, "quiz")
        self.assertEqual(intent.card_id, card_id)
        action.click()
        self.assertEqual(launched[0].card_id, card_id)


class ReviewCalendarScheduleTests(ReviewScheduleTestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.app = QApplication.instance() or QApplication([])

    def test_schedule_workspace_lists_unscheduled_current_cards(self) -> None:
        card_id = self._card_id(name="Synthetic Current Card")

        rows = list_card_schedules(today="2026-08-26")

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["card_id"], card_id)
        self.assertEqual(rows[0]["state"], "unscheduled")

    def test_retired_historical_event_cannot_select_replacement_schedule(self) -> None:
        old_card_id = self._card_id(name="Synthetic Historical Identity")
        now = "2026-08-26T12:00:00+00:00"
        with db.get_connection() as conn:
            old_card = conn.execute(
                "SELECT collection_id, card_number FROM cards WHERE id = ?",
                (old_card_id,),
            ).fetchone()
            entry_id = int(
                conn.execute(
                    """
                    SELECT membership.entry_id
                    FROM card_revisions AS revision
                    JOIN card_revision_entries AS membership
                      ON membership.revision_id = revision.id
                    WHERE revision.card_id = ?
                    ORDER BY revision.revision_number DESC
                    LIMIT 1
                    """,
                    (old_card_id,),
                ).fetchone()[0]
            )
        session_id = quiz.create_quiz_session(
            int(old_card["collection_id"]),
            int(old_card["card_number"]),
            "term_to_meaning",
            1,
        )
        quiz.record_quiz_answer(
            session_id,
            entry_id,
            "term",
            "meaning",
            "meaning",
            True,
        )
        quiz.complete_quiz_session(session_id)

        with db.get_connection() as conn:
            collection_id = int(old_card["collection_id"])
            conn.execute(
                "DELETE FROM entry_collections WHERE collection_id = ?",
                (collection_id,),
            )
            reconcile_collection_card_history(
                conn,
                collection_id,
                change_reason="synthetic_historical_retire",
            )
            replacement_entry_id = int(
                conn.execute(
                    """
                    INSERT INTO entries (
                        language, explanation_language, entry_type, term, meaning,
                        example, notes, tags, source, status, created_at, updated_at
                    ) VALUES ('English', 'English', 'word', 'replacement', 'new meaning',
                              '', '', '', '', 'new', ?, ?)
                    """,
                    (now, now),
                ).lastrowid
            )
            conn.execute(
                """
                INSERT INTO entry_collections (entry_id, collection_id, position, added_at)
                VALUES (?, ?, 1, ?)
                """,
                (replacement_entry_id, collection_id, now),
            )
            reconcile_collection_card_history(
                conn,
                collection_id,
                change_reason="synthetic_historical_replacement",
            )
            replacement_card_id = int(
                conn.execute(
                    """
                    SELECT id FROM cards
                    WHERE collection_id = ? AND card_number = 1 AND is_active = 1
                    """,
                    (collection_id,),
                ).fetchone()[0]
            )

        controller = ReviewCalendarController()
        controller.select_card_event(
            card_id=old_card_id,
            collection_id=collection_id,
            card_number=1,
            collection_name="Synthetic Historical Identity",
        )

        self.assertEqual(controller.selected_card_id, old_card_id)
        self.assertEqual(len(controller.card_history), 1)
        self.assertEqual(controller.card_history[0]["card_id"], old_card_id)
        self.assertIsNone(controller.current_schedule)
        with self.assertRaises(ValueError):
            controller.set_selected_next_review("2026-08-30", today="2026-08-26")
        self.assertEqual(
            get_card_schedule(replacement_card_id, today="2026-08-26")["state"],
            "unscheduled",
        )

    def test_selected_card_schedule_can_be_read_and_edited_separately(self) -> None:
        card_id = self._card_id(name="Synthetic Calendar Schedule")
        with db.get_connection() as conn:
            card = conn.execute(
                "SELECT collection_id, card_number FROM cards WHERE id = ?",
                (card_id,),
            ).fetchone()
        controller = ReviewCalendarController()

        controller.select_card(
            int(card["collection_id"]),
            int(card["card_number"]),
            "Synthetic Calendar Schedule",
        )
        self.assertEqual(controller.current_schedule["state"], "unscheduled")

        controller.set_selected_next_review("2026-08-29", today="2026-08-26")

        self.assertEqual(controller.current_schedule["card_id"], card_id)
        self.assertEqual(controller.current_schedule["next_due_at"], "2026-08-29")
        self.assertEqual(controller.card_history, [])
        self.assertEqual(controller.legacy_logs, [])

    def test_schedule_table_and_date_editor_update_the_selected_card(self) -> None:
        card_id = self._card_id(name="Synthetic Schedule UI")
        set_card_next_review(card_id, date.today().isoformat())
        controller = ReviewCalendarController()
        view = ReviewCalendarView(controller)
        self.addCleanup(view.deleteLater)

        view.refresh()
        view._schedule_table.selectRow(0)
        view._schedule_date.setDate(QDate(2026, 8, 30))
        view._schedule_save_button.click()

        self.assertEqual(controller.current_schedule["card_id"], card_id)
        self.assertEqual(controller.current_schedule["next_due_at"], "2026-08-30")
        self.assertEqual(view._history_table.rowCount(), 0)

        with patch.object(QMessageBox, "question", return_value=QMessageBox.StandardButton.Yes):
            view._schedule_clear_button.click()

        self.assertEqual(controller.current_schedule["state"], "unscheduled")
        self.assertEqual(
            list_actionable_schedules(today="2026-08-30"),
            [],
        )


if __name__ == "__main__":
    unittest.main()
